# -*- coding: utf-8 -*-
import os
import csv
from collections import defaultdict
import pandas as pd

def read_file(path: str) -> list:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [line.rstrip("\r\n") for line in f]

def parse_headers(lines: list) -> dict:
    h = {}
    for line in lines:
        if line.startswith("H"):
            code = line[:3].strip()
            h[code] = line[3:].strip()
    return h

def parse_structure(lines: list) -> dict:
    """
    Devuelve estructura jerarquizada:
      {
        disparo: {
          linea: [(estaca_ini, estaca_fin), ...]
        }
      }
    También devuelve set de disparos y dict de líneas por disparo.
    """
    structure = defaultdict(lambda: defaultdict(list))
    for line in lines:
        if not line.startswith("X"):
            continue
        fields = line.split()
        if len(fields) < 10:
            continue
        try:
            disparo     = fields[3]
            linea       = fields[7]
            estaca_ini  = fields[8]
            estaca_fin  = fields[9]
        except IndexError:
            continue
        structure[disparo][linea].append((estaca_ini, estaca_fin))
    return structure

def compare_structures(s1: dict, s2: dict, name1: str, name2: str) -> dict:
    """
    Compara las dos estructuras jerarquizadas.
    Devuelve un dict con todos los resultados organizados.
    """
    def get_sort_key(x):
        try:
            return float(x.lstrip('-'))
        except ValueError:
            return x

    all_disparos = sorted(set(s1.keys()) | set(s2.keys()), key=get_sort_key)

    results = []

    for disp in all_disparos:
        in1 = disp in s1
        in2 = disp in s2

        if not in1:
            results.append({
                "disparo": disp,
                "status": "only2",
                "lineas": [],
            })
            continue
        if not in2:
            results.append({
                "disparo": disp,
                "status": "only1",
                "lineas": [],
            })
            continue

        # Disparo existe en ambos — comparar líneas
        lineas1 = s1[disp]
        lineas2 = s2[disp]
        all_lineas = sorted(set(lineas1.keys()) | set(lineas2.keys()), key=get_sort_key)

        linea_results = []
        disp_status = "ok"

        for lin in all_lineas:
            l_in1 = lin in lineas1
            l_in2 = lin in lineas2

            if not l_in1:
                linea_results.append({
                    "linea": lin, "status": "only2",
                    "estacas1": [], "estacas2": lineas2[lin],
                    "estaca_diffs": []
                })
                disp_status = "diff"
                continue
            if not l_in2:
                linea_results.append({
                    "linea": lin, "status": "only1",
                    "estacas1": lineas1[lin], "estacas2": [],
                    "estaca_diffs": []
                })
                disp_status = "diff"
                continue

            # Línea en ambos — comparar estacas
            set1 = set(lineas1[lin])
            set2 = set(lineas2[lin])

            estaca_diffs = []
            only_in_1 = sorted(set1 - set2, key=lambda x: get_sort_key(x[0]))
            only_in_2 = sorted(set2 - set1, key=lambda x: get_sort_key(x[0]))

            for e in only_in_1:
                estaca_diffs.append({"tipo": "solo_arch1", "ini": e[0], "fin": e[1]})
            for e in only_in_2:
                estaca_diffs.append({"tipo": "solo_arch2", "ini": e[0], "fin": e[1]})

            lin_status = "diff" if estaca_diffs else "ok"
            if lin_status == "diff":
                disp_status = "diff"

            linea_results.append({
                "linea": lin,
                "status": lin_status,
                "estacas1": sorted(lineas1[lin], key=lambda x: get_sort_key(x[0])),
                "estacas2": sorted(lineas2[lin], key=lambda x: get_sort_key(x[0])),
                "estaca_diffs": estaca_diffs,
            })

        results.append({
            "disparo": disp,
            "status": disp_status,
            "n_lineas1": len(lineas1),
            "n_lineas2": len(lineas2),
            "lineas": linea_results,
        })

    disparos_solo1   = [r for r in results if r["status"] == "only1"]
    disparos_solo2   = [r for r in results if r["status"] == "only2"]
    disparos_diff    = [r for r in results if r["status"] == "diff"]
    disparos_ok      = [r for r in results if r["status"] == "ok"]

    return {
        "results": results,
        "n_disparos1": len(s1),
        "n_disparos2": len(s2),
        "disparos_solo1": disparos_solo1,
        "disparos_solo2": disparos_solo2,
        "disparos_diff":  disparos_diff,
        "disparos_ok":    disparos_ok,
        "name1": name1,
        "name2": name2,
    }

def run_comparison(path1: str, path2: str) -> dict:
    lines1 = read_file(path1)
    lines2 = read_file(path2)
    h1, h2 = parse_headers(lines1), parse_headers(lines2)
    s1 = parse_structure(lines1)
    s2 = parse_structure(lines2)
    cmp = compare_structures(s1, s2,
                             os.path.basename(path1),
                             os.path.basename(path2))
    cmp["h_diffs"] = []
    for k in sorted(set(h1) | set(h2)):
        v1 = h1.get(k, "<ausente>")
        v2 = h2.get(k, "<ausente>")
        if v1 != v2:
            cmp["h_diffs"].append({"code": k, "v1": v1, "v2": v2})
    return cmp

def export_txt(result: dict, path: str):
    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    with open(path, "w", encoding="utf-8") as f:
        f.write("REPORTE DE DIFERENCIAS SPS/XPS\n")
        f.write(f"Archivo 1 : {result['name1']}  ({n1} disparos)\n")
        f.write(f"Archivo 2 : {result['name2']}  ({n2} disparos)\n")
        f.write("=" * 72 + "\n\n")

        if n1 != n2:
            f.write(f"⚠ DIFERENCIA EN DISPAROS: Archivo 1={n1}, Arch2={n2}\n\n")

        if result["h_diffs"]:
            f.write("CABECERAS DISTINTAS:\n")
            for d in result["h_diffs"]:
                f.write(f"  [{d['code']}]  Archivo 1: {d['v1']}  →  Archivo 2: {d['v2']}\n")
            f.write("\n")

        if result["disparos_solo1"]:
            f.write("DISPAROS SOLO EN ARCHIVO 1:\n")
            for r in result["disparos_solo1"]:
                f.write(f"  Disparo {r['disparo']}\n")
            f.write("\n")

        if result["disparos_solo2"]:
            f.write("DISPAROS SOLO EN ARCHIVO 2:\n")
            for r in result["disparos_solo2"]:
                f.write(f"  Disparo {r['disparo']}\n")
            f.write("\n")

        for r in result["results"]:
            if r["status"] == "ok":
                continue
            if r["status"] in ("only1", "only2"):
                arch = "Archivo 1" if r["status"] == "only1" else "Archivo 2"
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en {arch}\n\n")
                continue

            f.write(f"DISPARO {r['disparo']}\n")
            if r.get("n_lineas1") != r.get("n_lineas2"):
                f.write(f"  ⚠ Líneas: Arch1={r['n_lineas1']}, Arch2={r['n_lineas2']}\n")

            for lr in r["lineas"]:
                if lr["status"] == "ok":
                    continue
                if lr["status"] == "only1":
                    f.write(f"  Línea {lr['linea']}  ⚠ Solo en Archivo 1\n")
                    for e in lr["estacas1"]:
                        f.write(f"    Estacas {e[0]} → {e[1]}\n")
                elif lr["status"] == "only2":
                    f.write(f"  Línea {lr['linea']}  ⚠ Solo en Archivo 2\n")
                    for e in lr["estacas2"]:
                        f.write(f"    Estacas {e[0]} → {e[1]}\n")
                else:
                    f.write(f"  Línea {lr['linea']}  ({len(lr['estaca_diffs'])} diferencia(s) en estacas)\n")
                    for ed in lr["estaca_diffs"]:
                        arch = "Archivo 1" if ed["tipo"] == "solo_arch1" else "Archivo 2"
                        f.write(f"    Solo en {arch}: estaca {ed['ini']} → {ed['fin']}\n")
            f.write("\n")

def export_csv(result: dict, path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["disparo", "linea", "tipo_diferencia",
                    "archivo", "estaca_ini", "estaca_fin"])
        for r in result["results"]:
            if r["status"] == "ok":
                continue
            if r["status"] == "only1":
                w.writerow([r["disparo"], "—", "disparo_faltante",
                             "Archivo 2", "—", "—"])
                continue
            if r["status"] == "only2":
                w.writerow([r["disparo"], "—", "disparo_faltante",
                             "Archivo 1", "—", "—"])
                continue
            for lr in r["lineas"]:
                if lr["status"] == "ok":
                    continue
                if lr["status"] == "only1":
                    for e in lr["estacas1"]:
                        w.writerow([r["disparo"], lr["linea"], "linea_faltante",
                                    "Archivo 2", e[0], e[1]])
                elif lr["status"] == "only2":
                    for e in lr["estacas2"]:
                        w.writerow([r["disparo"], lr["linea"], "linea_faltante",
                                    "Archivo 1", e[0], e[1]])
                else:
                    for ed in lr["estaca_diffs"]:
                        arch = "Archivo 1" if ed["tipo"] == "solo_arch1" else "Archivo 2"
                        w.writerow([r["disparo"], lr["linea"], "estaca_faltante",
                                    arch, ed["ini"], ed["fin"]])

def export_comparativa(result: dict, path: str):
    """
    Exporta la comparativa completa: todos los disparos, todas las líneas
    y todas las estacas — tanto coincidentes como diferentes.
    """
    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    W = 76

    def sep(ch="─"): return ch * W

    with open(path, "w", encoding="utf-8") as f:
        f.write("COMPARATIVA COMPLETA SPS/XPS\n")
        f.write(f"Archivo 1 : {result['name1']}  ({n1} disparos)\n")
        f.write(f"Archivo 2 : {result['name2']}  ({n2} disparos)\n")
        f.write(sep("═") + "\n\n")

        # Resumen ejecutivo
        ndiff  = len(result["disparos_diff"])
        nmiss1 = len(result["disparos_solo1"])
        nmiss2 = len(result["disparos_solo2"])
        nok    = len(result["disparos_ok"])
        f.write("RESUMEN\n")
        f.write(f"  Disparos totales comparados : {len(result['results'])}\n")
        f.write(f"  Idénticos                   : {nok}\n")
        f.write(f"  Con diferencias de líneas/estacas : {ndiff}\n")
        f.write(f"  Solo en Archivo 1           : {nmiss1}\n")
        f.write(f"  Solo en Archivo 2           : {nmiss2}\n")
        if n1 != n2:
            diff = abs(n1 - n2)
            which = "Archivo 1" if n1 > n2 else "Archivo 2"
            f.write(f"  ⚠ {which} tiene {diff} disparo(s) extra\n")
        f.write("\n")

        if result["h_diffs"]:
            f.write("DIFERENCIAS EN CABECERAS (H)\n")
            f.write(sep() + "\n")
            for d in result["h_diffs"]:
                f.write(f"  [{d['code']}]\n")
                f.write(f"    Archivo 1: {d['v1']}\n")
                f.write(f"    Archivo 2: {d['v2']}\n")
            f.write("\n")

        # Detalle por disparo
        for r in result["results"]:
            f.write(sep("═") + "\n")
            if r["status"] == "only1":
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en Archivo 1\n\n")
                continue
            if r["status"] == "only2":
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en Archivo 2\n\n")
                continue

            status_lbl = "✔ Idéntico" if r["status"] == "ok" else "⚠ Con diferencias"
            f.write(f"DISPARO {r['disparo']}  [{status_lbl}]  "
                    f"(Arch1: {r.get('n_lineas1','?')} líneas  |  "
                    f"Arch2: {r.get('n_lineas2','?')} líneas)\n")
            f.write(sep() + "\n")

            # Encabezado de tabla
            f.write(f"  {'Línea':<10} {'Estado':<22} {'Estacas Archivo 1':<22} "
                    f"{'Estacas Archivo 2':<22} Diferencias\n")
            f.write(f"  {'-' * (W-2)}\n")

            for lr in r["lineas"]:
                s = lr["status"]

                def fmt(lst):
                    if not lst: return "—"
                    if len(lst) == 1: return f"{lst[0][0]}→{lst[0][1]}"
                    def _n(x):
                        try: return float(x)
                        except: return 0
                    ini = min((e[0] for e in lst), key=_n)
                    fin = max((e[1] for e in lst), key=_n)
                    return f"{ini}→{fin} ({len(lst)})"

                if s == "only1":
                    estado  = "Solo en Archivo 1"
                    e1_txt  = fmt(lr["estacas1"])
                    e2_txt  = "—"
                    diff_tx = ""
                elif s == "only2":
                    estado  = "Solo en Archivo 2"
                    e1_txt  = "—"
                    e2_txt  = fmt(lr["estacas2"])
                    diff_tx = ""
                else:
                    estado  = "✔ ok" if s == "ok" else f"⚠ {len(lr['estaca_diffs'])} dif."
                    e1_txt  = fmt(lr["estacas1"])
                    e2_txt  = fmt(lr["estacas2"])
                    parts   = []
                    for ed in lr["estaca_diffs"]:
                        a = "Arch1" if ed["tipo"] == "solo_arch1" else "Arch2"
                        parts.append(f"{a}:{ed['ini']}→{ed['fin']}")
                    diff_tx = " | ".join(parts)

                f.write(f"  {lr['linea']:<10} {estado:<22} {e1_txt:<22} "
                        f"{e2_txt:<22} {diff_tx}\n")

            f.write("\n")

def export_xlsx(result: dict, path: str):
    """
    Exporta la comparativa completa a Excel (.xlsx) con formato visual.
    3 hojas: Resumen · Comparativa completa · Solo diferencias
    Requiere: openpyxl  (pip install openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "Se necesita openpyxl para exportar a Excel.\n"
            "Instálalo con:  pip install openpyxl")

    HEX = {
        "ok_bg":    "C8EDD6", "ok_fg":    "1B5E35",
        "diff_bg":  "FDECC8", "diff_fg":  "7B4D00",
        "only1_bg": "D0E8FB", "only1_fg": "0C3D6E",
        "only2_bg": "E8DFFB", "only2_fg": "3D1E87",
        "hdr_bg":   "2C3E50", "hdr_fg":   "FFFFFF",
        "sub_bg":   "546E7A", "sub_fg":   "FFFFFF",
        "alt_bg":   "F7F6F3", "white":    "FFFFFF",
        "border":   "BDB8AD",
        "miss_bg":  "FFB347", "miss_fg": "7A3500",
    }

    STATUS_STYLE = {
        "ok":    ("ok_bg",    "ok_fg"),
        "diff":  ("diff_bg",  "diff_fg"),
        "only1": ("only1_bg", "only1_fg"),
        "only2": ("only2_bg", "only2_fg"),
    }
    LINE_MISS_STYLE = ("miss_bg", "miss_fg")   # naranja para líneas faltantes
    STATUS_LABEL = {
        "ok":    "✔ Idéntico",
        "diff":  "⚠ Con diferencias",
        "only1": "✖ Solo en Archivo 1",
        "only2": "✖ Solo en Archivo 2",
    }

    _side = Side(style="thin", color=HEX["border"])
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    def _fill(h):  return PatternFill("solid", fgColor=h)
    def _font(h, bold=False, sz=10):
        return Font(color=h, bold=bold, size=sz, name="Calibri")
    def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def apply(cell, bg=None, fg="1A1917", bold=False, align=None, sz=10):
        if bg: cell.fill = _fill(bg)
        cell.font      = _font(fg, bold=bold, sz=sz)
        cell.alignment = align or _left()
        cell.border    = _border

    def fmt_estacas(lst):
        if not lst: return "—"
        def _n(x):
            try: return float(x)
            except: return 0
        ini = min((e[0] for e in lst), key=_n)
        fin = max((e[1] for e in lst), key=_n)
        return f"{ini} → {fin}" if len(lst) == 1 else f"{ini} → {fin}  ({len(lst)} rangos)"

    HEADERS = ["DISPARO", "ESTADO DISPARO", "LÍNEA",
               "ESTACAS ARCHIVO 1", "ESTACAS ARCHIVO 2",
               "ESTADO LÍNEA", "DIFERENCIAS DE ESTACAS"]

    _col_max = {}   # sheet_title -> {col_idx: max_len}

    def _track(ws_title, col, value):
        length = len(str(value)) if value else 0
        if ws_title not in _col_max:
            _col_max[ws_title] = {}
        if col not in _col_max[ws_title] or length > _col_max[ws_title][col]:
            _col_max[ws_title][col] = length

    def auto_fit(ws):
        col_lens = _col_max.get(ws.title, {})
        for col, max_len in col_lens.items():
            width = min(max(max_len + 4, 10), 60)
            ws.column_dimensions[get_column_letter(col)].width = width

    def setup_sheet(ws):
        ws.row_dimensions[1].height = 24
        ws.append(HEADERS)
        for col, h in enumerate(HEADERS, 1):
            apply(ws.cell(1, col), bg=HEX["hdr_bg"], fg=HEX["hdr_fg"],
                  bold=True, align=_center(), sz=10)
            _track(ws.title, col, h)
        ws.freeze_panes = "A2"

    def write_shot_rows(ws, shot_list, only_diffs=False):
        alt = False
        for shot in shot_list:
            s = shot["status"]
            if only_diffs and s == "ok":
                continue
            bg_k, fg_k = STATUS_STYLE.get(s, ("alt_bg", "1A1917"))
            d_bg, d_fg = HEX[bg_k], HEX[fg_k]
            d_lbl = STATUS_LABEL.get(s, s)

            if s in ("only1", "only2"):
                row_vals = [shot["disparo"], d_lbl, "—", "—", "—", "—", ""]
                ws.append(row_vals)
                r = ws.max_row
                ws.row_dimensions[r].height = 18
                for col, val in enumerate(row_vals, 1):
                    apply(ws.cell(r, col), bg=d_bg, fg=d_fg,
                          bold=(col <= 2), align=_center())
                    _track(ws.title, col, val)
                continue

            for lr in shot["lineas"]:
                ls = lr["status"]
                if only_diffs and ls == "ok":
                    continue
                if ls in ("only1", "only2"):
                    ls_bk, ls_fk = LINE_MISS_STYLE
                else:
                    ls_bk, ls_fk = STATUS_STYLE.get(ls, ("alt_bg", "1A1917"))
                l_bg, l_fg = HEX[ls_bk], HEX[ls_fk]
                l_lbl = STATUS_LABEL.get(ls, ls)
                e1 = fmt_estacas(lr["estacas1"]) if lr["estacas1"] else "—"
                e2 = fmt_estacas(lr["estacas2"]) if lr["estacas2"] else "—"
                parts = []
                for ed in lr["estaca_diffs"]:
                    a = "Arch.1" if ed["tipo"] == "solo_arch1" else "Arch.2"
                    parts.append(f"Solo {a}: {ed['ini']}→{ed['fin']}")
                diffs_txt = "  |  ".join(parts)

                alt = not alt
                row_bg = HEX["alt_bg"] if alt else HEX["white"]
                row_vals = [shot["disparo"], d_lbl, lr["linea"],
                            e1, e2, l_lbl, diffs_txt]

                ws.append(row_vals)
                r = ws.max_row
                ws.row_dimensions[r].height = 18

                for col, val in enumerate(row_vals, 1):
                    _track(ws.title, col, val)

                if ls in ("only1", "only2"):
                    for col in range(1, 8):
                        apply(ws.cell(r, col), bg=l_bg, fg=l_fg,
                              bold=(col <= 3), align=_center())
                    apply(ws.cell(r, 7), bg=l_bg, fg=l_fg, align=_left())
                else:
                    apply(ws.cell(r, 1), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 2), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 3), bg=l_bg,    fg=l_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 4), bg=row_bg,                                     align=_center())
                    apply(ws.cell(r, 5), bg=row_bg,                                     align=_center())
                    apply(ws.cell(r, 6), bg=l_bg,    fg=l_fg,        bold=(ls != "ok"), align=_center())
                    d_cell_bg = HEX["diff_bg"] if diffs_txt else row_bg
                    d_cell_fg = HEX["diff_fg"] if diffs_txt else "1A1917"
                    apply(ws.cell(r, 7), bg=d_cell_bg, fg=d_cell_fg,
                          bold=bool(diffs_txt), align=_left())

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Resumen"
    ws_r.column_dimensions["A"].width = 38
    ws_r.column_dimensions["B"].width = 18
    ws_r.column_dimensions["C"].width = 28

    ws_r.merge_cells("A1:C1")
    apply(ws_r["A1"], bg=HEX["hdr_bg"], fg=HEX["hdr_fg"],
          bold=True, align=_center(), sz=14)
    ws_r["A1"].value = "COMPARATIVA SPS / XPS"
    ws_r.row_dimensions[1].height = 36

    for i, (lbl, val) in enumerate([
        ("Archivo 1", result["name1"]),
        ("Archivo 2", result["name2"]),
    ], start=2):
        ws_r.merge_cells(f"A{i}:C{i}")
        c = ws_r.cell(i, 1, f"  {lbl}:   {val}")
        apply(c, bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True)
        ws_r.row_dimensions[i].height = 20
        _track(ws_r.title, 1, f"  {lbl}:   {val}")

    ws_r.append([])
    ws_r.append(["INDICADOR", "VALOR", "NOTA"])
    for col in range(1, 4):
        apply(ws_r.cell(ws_r.max_row, col),
              bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True, align=_center())
    ws_r.row_dimensions[ws_r.max_row].height = 20

    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    ndiff  = len(result["disparos_diff"])
    nmiss1 = len(result["disparos_solo1"])
    nmiss2 = len(result["disparos_solo2"])
    nok    = len(result["disparos_ok"])

    rows_sum = [
        ("Disparos en Archivo 1",             n1,     ""),
        ("Disparos en Archivo 2",             n2,
         "⚠ Cantidades distintas" if n1 != n2 else "✔ Coinciden"),
        ("Disparos idénticos",                nok,    ""),
        ("Disparos con difs. líneas/estacas", ndiff,  ""),
        ("Solo en Archivo 1",                 nmiss1, ""),
        ("Solo en Archivo 2",                 nmiss2, ""),
    ]
    for i, (lbl, val, nota) in enumerate(rows_sum):
        ws_r.append([lbl, val, nota])
        r = ws_r.max_row
        row_bg = HEX["alt_bg"] if i % 2 == 0 else HEX["white"]
        for col, v in enumerate([lbl, val, nota], 1):
            apply(ws_r.cell(r, col), bg=row_bg)
            _track(ws_r.title, col, v)
        if "difs" in lbl and val > 0:
            apply(ws_r.cell(r, 2), bg=HEX["diff_bg"], fg=HEX["diff_fg"], bold=True, align=_center())
        elif "Solo en" in lbl and val > 0:
            apply(ws_r.cell(r, 2), bg=HEX["only1_bg"], fg=HEX["only1_fg"], bold=True, align=_center())
        elif lbl == "Disparos idénticos" and val == n1 == n2:
            apply(ws_r.cell(r, 2), bg=HEX["ok_bg"], fg=HEX["ok_fg"], bold=True, align=_center())
        ws_r.row_dimensions[r].height = 18

    # ── Hoja 2: Comparativa completa ─────────────────────────────────────
    ws_full = wb.create_sheet("Comparativa completa")
    setup_sheet(ws_full)
    write_shot_rows(ws_full, result["results"], only_diffs=False)
    auto_fit(ws_full)

    # ── Hoja 3: Solo diferencias ─────────────────────────────────────────
    ws_diff = wb.create_sheet("Solo diferencias")
    setup_sheet(ws_diff)
    write_shot_rows(ws_diff, result["results"], only_diffs=True)
    auto_fit(ws_diff)

    auto_fit(ws_r)
    wb.save(path)
