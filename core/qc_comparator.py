# -*- coding: utf-8 -*-

import pandas as pd
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core.lectores import leer_sps_rps, detectar_tipo_sps_rps, ENCABEZADOS_RPS, ENCABEZADOS_SPS, ENCABEZADOS_XPS, COLUMNAS_RPS, COLUMNAS_XPS, leer_xps_completo
from core.comparador_disparos import run_comparison

def aplicar_estilo_hoja(ws, color_encabezado, es_diferencias=False):
    """Aplica diseño estético y legibilidad a la hoja usando openpyxl."""
    # Relleno del encabezado (color sólido)
    fill_encabezado = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type='solid')
    font_encabezado = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    align_encabezado = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Formatear la primera fila (encabezado)
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_encabezado
        cell.font = font_encabezado
        cell.alignment = align_encabezado
        
    # Si la hoja tiene más de 15,000 filas, desactivamos el estilizado celda por celda
    # para evitar degradación de rendimiento y consumo masivo de memoria.
    limite_filas = 15000
    
    if ws.max_row <= limite_filas:
        # Relleno de filas alternas (cebreado muy tenue)
        fill_cebra = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        # Bordes inferiores finos para separar filas
        borde_inferior = Border(bottom=Side(style='thin', color='E0E0E0'))
        
        # Estilos para celdas de datos
        font_datos = Font(name='Segoe UI', size=10)
        
        # Formatear las filas de datos
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            usar_cebra = (row_idx % 2 == 0)
            
            # Si es la hoja de diferencias, pintamos según el Tipo de Cambio
            fill_fila_dif = None
            if es_diferencias:
                tipo_cambio_val = None
                for c_idx in range(1, ws.max_column + 1):
                    header_val = ws.cell(row=1, column=c_idx).value
                    if header_val == 'Tipo de Cambio':
                        tipo_cambio_val = ws.cell(row=row_idx, column=c_idx).value
                        break
                
                if tipo_cambio_val == 'Eliminado':
                    fill_fila_dif = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')  # Durazno claro
                elif tipo_cambio_val == 'Nuevo':
                    fill_fila_dif = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # Verde claro
                elif tipo_cambio_val == 'Cambio de Coordenada':
                    fill_fila_dif = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Amarillo claro
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_datos
                cell.border = borde_inferior
                
                # Relleno
                if fill_fila_dif:
                    cell.fill = fill_fila_dif
                elif usar_cebra:
                    cell.fill = fill_cebra
                    
                # Alineación y formatos numéricos basados en el nombre de columna
                header_val = ws.cell(row=1, column=col_idx).value
                
                if header_val:
                    h_lower = str(header_val).lower()
                    if any(x in h_lower for x in ['coordenada x', 'x base', 'x nuevo', 'coordenada y', 'y base', 'y nuevo']):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '0.0'
                    elif any(x in h_lower for x in ['elevación', 'elevacion']):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '0.0'
                    elif any(x in h_lower for x in ['línea', 'linea', 'punto', 'desde', 'hasta', 'incremento', 'evento', 'reel', 'carrete', 'estático', 'estatico']):
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.number_format = '0'
                    elif any(x in h_lower for x in ['tipo de cambio', 'paquete']):
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
    # Ajustar anchos de columna dinámicamente usando una muestra de las primeras 200 filas
    filas_muestra = min(ws.max_row, 200)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, filas_muestra + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
        
    # Congelar fila superior
    ws.freeze_panes = 'A2'
    
    # Forzar cuadrícula visible
    ws.sheet_view.showGridLines = True

def comparar_carpetas(lista_carpetas, progreso=None):
    """
    Compara paquetes diarios a partir de una lista de carpetas.
    Cada carpeta debe contener archivos RPS, SPS y/o XPS.
    Retorna un dict con las diferencias y también guarda los datos originales en el Excel.
    """
    if progreso:
        progreso.actualizar(5, "Analizando carpetas...")
    
    carpetas_ordenadas = sorted(lista_carpetas, key=lambda x: os.path.basename(x))
    
    paquetes = {}
    for carpeta in carpetas_ordenadas:
        nombre = os.path.basename(carpeta)
        archivos = os.listdir(carpeta)
        rps_path = None
        sps_path = None
        xps_path = None
        
        # Primero buscar por extensión conocida
        for archivo in archivos:
            ruta_completa = os.path.join(carpeta, archivo)
            ext = os.path.splitext(archivo)[1].lower()
            if ext in ['.rps', '.rcp']:
                rps_path = ruta_completa
            elif ext == '.sps':
                sps_path = ruta_completa
            elif ext == '.xps':
                xps_path = ruta_completa
        
        # Si no se encontraron por extensión, buscar por contenido (archivos .txt o sin extensión)
        if not rps_path and not sps_path and not xps_path:
            for archivo in archivos:
                ruta_completa = os.path.join(carpeta, archivo)
                ext = os.path.splitext(archivo)[1].lower()
                if ext in ['.txt', '']:
                    tipo = detectar_tipo_sps_rps(ruta_completa)
                    if tipo == 'R' and not rps_path:
                        rps_path = ruta_completa
                    elif tipo == 'S' and not sps_path:
                        sps_path = ruta_completa
        
        # Leer archivos (siempre devuelven DataFrames con columnas estándar)
        df_rps = leer_sps_rps(rps_path, 'R') if rps_path else pd.DataFrame(columns=COLUMNAS_RPS)
        df_sps = leer_sps_rps(sps_path, 'S') if sps_path else pd.DataFrame(columns=COLUMNAS_RPS)
        
        # Leer XPS usando slicing exacto
        df_xps = leer_xps_completo(xps_path) if xps_path else pd.DataFrame(columns=COLUMNAS_XPS)
        
        paquetes[nombre] = {
            'carpeta': carpeta,
            'rps': df_rps,
            'sps': df_sps,
            'xps': df_xps,
            'archivos': {
                'rps': os.path.basename(rps_path) if rps_path else None,
                'sps': os.path.basename(sps_path) if sps_path else None,
                'xps': os.path.basename(xps_path) if xps_path else None
            },
            'rutas': {
                'rps': rps_path,
                'sps': sps_path,
                'xps': xps_path
            }
        }
    
    if progreso:
        progreso.actualizar(30, f"Datos cargados de {len(paquetes)} paquetes.")
        for nombre, datos in paquetes.items():
            progreso.actualizar(32, f"  {nombre}: RPS={len(datos['rps'])}, SPS={len(datos['sps'])}, XPS={len(datos['xps'])}")
    
    # Comparar paquetes consecutivos
    diferencias = {'RPS': [], 'SPS': [], 'XPS': []}
    xps_comparaciones = []
    
    for i in range(len(carpetas_ordenadas)-1):
        nombre_base = carpetas_ordenadas[i]
        nombre_nuevo = carpetas_ordenadas[i+1]
        base_name = os.path.basename(nombre_base)
        new_name = os.path.basename(nombre_nuevo)
        
        if progreso:
            progreso.actualizar(35 + int(i/(len(carpetas_ordenadas)-1)*40), 
                              f"Comparando {base_name} vs {new_name}")
        
        # RPS
        df_base = paquetes[base_name]['rps']
        df_nuevo = paquetes[new_name]['rps']
        if not df_base.empty or not df_nuevo.empty:
            df_diff = comparar_sin_merge(df_base, df_nuevo, ['linea', 'punto'], ['x', 'y', 'elevacion'])
            if not df_diff.empty:
                df_diff.insert(0, 'Paquete Base', base_name)
                df_diff.insert(1, 'Paquete Nuevo', new_name)
                diferencias['RPS'].append(df_diff)
        
        # SPS
        df_base = paquetes[base_name]['sps']
        df_nuevo = paquetes[new_name]['sps']
        if not df_base.empty or not df_nuevo.empty:
            df_diff = comparar_sin_merge(df_base, df_nuevo, ['linea', 'punto'], ['x', 'y', 'elevacion'])
            if not df_diff.empty:
                df_diff.insert(0, 'Paquete Base', base_name)
                df_diff.insert(1, 'Paquete Nuevo', new_name)
                diferencias['SPS'].append(df_diff)
        
        # XPS (Lógica superior del comparador)
        xps_path_base = paquetes[base_name]['rutas']['xps']
        xps_path_nuevo = paquetes[new_name]['rutas']['xps']
        if xps_path_base and xps_path_nuevo:
            try:
                cmp_res = run_comparison(xps_path_base, xps_path_nuevo)
                xps_comparaciones.append({
                    'par': f"{base_name} vs {new_name}",
                    'resultado': cmp_res
                })
            except Exception as e:
                print(f"Error al comparar XPS {base_name} vs {new_name}: {e}")
                
        # Mantener diferencias XPS antiguas para compatibilidad interna si fuera necesario,
        # pero ahora la exportación a Excel usará la lógica superior.
        df_base = paquetes[base_name]['xps']
        df_nuevo = paquetes[new_name]['xps']
        if not df_base.empty or not df_nuevo.empty:
            df_diff = comparar_sin_merge(df_base, df_nuevo, 
                                             ['linea_f', 'punto_f', 'linea_r', 'desde', 'hasta'], None)
            if not df_diff.empty:
                df_diff.insert(0, 'Paquete Base', base_name)
                df_diff.insert(1, 'Paquete Nuevo', new_name)
                diferencias['XPS'].append(df_diff)
    
    # Combinar diferencias
    diferencias_final = {}
    for tipo in ['RPS', 'SPS', 'XPS']:
        if diferencias[tipo]:
            diferencias_final[tipo] = pd.concat(diferencias[tipo], ignore_index=True)
    
    if progreso:
        progreso.actualizar(90, "Generando Excel con datos originales y diferencias...")
    
    # Exportar a Excel incluyendo datos originales con encabezados descriptivos y comparativas XPS
    exportar_excel_completo(paquetes, diferencias_final, xps_comparaciones)
    
    return {'diferencias': diferencias_final, 'paquetes': paquetes, 'xps_comparaciones': xps_comparaciones}

def comparar_sin_merge(df_base, df_nuevo, key_cols, coord_cols=None):
    """
    Compara dos DataFrames sin usar pandas.merge (solo diccionarios).
    """
    if df_base.empty and df_nuevo.empty:
        return pd.DataFrame()
    
    # Asegurar columnas clave
    for col in key_cols:
        if col not in df_base.columns:
            df_base[col] = ''
        if col not in df_nuevo.columns:
            df_nuevo[col] = ''
    
    # Convertir a listas de diccionarios
    base_records = df_base.to_dict('records')
    new_records = df_nuevo.to_dict('records')
    
    # Crear diccionarios clave -> fila completa
    base_dict = {}
    for rec in base_records:
        key = tuple(str(rec.get(col, '')) for col in key_cols)
        base_dict[key] = rec
    
    new_dict = {}
    for rec in new_records:
        key = tuple(str(rec.get(col, '')) for col in key_cols)
        new_dict[key] = rec
    
    base_keys = set(base_dict.keys())
    new_keys = set(new_dict.keys())
    
    eliminados = base_keys - new_keys
    nuevos = new_keys - base_keys
    comunes = base_keys & new_keys
    
    diferencias = []
    
    # Procesar eliminados
    for key in eliminados:
        fila = base_dict[key]
        dif = {col: fila.get(col, '') for col in key_cols}
        dif['Tipo de Cambio'] = 'Eliminado'
        if coord_cols:
            for col in coord_cols:
                if col in fila:
                    dif[col + '_base'] = fila[col]
        diferencias.append(dif)
    
    # Procesar nuevos
    for key in nuevos:
        fila = new_dict[key]
        dif = {col: fila.get(col, '') for col in key_cols}
        dif['Tipo de Cambio'] = 'Nuevo'
        if coord_cols:
            for col in coord_cols:
                if col in fila:
                    dif[col + '_nuevo'] = fila[col]
        diferencias.append(dif)
    
    # Procesar cambios de coordenadas
    if coord_cols and comunes:
        for key in comunes:
            fila_base = base_dict[key]
            fila_nuevo = new_dict[key]
            cambio = False
            for col in coord_cols:
                base_val = fila_base.get(col)
                nuevo_val = fila_nuevo.get(col)
                if pd.isna(base_val) and pd.isna(nuevo_val):
                    continue
                if base_val != nuevo_val:
                    cambio = True
                    break
            if cambio:
                dif = {col: fila_base.get(col, '') for col in key_cols}
                dif['Tipo de Cambio'] = 'Cambio de Coordenada'
                for col in coord_cols:
                    dif[col + '_base'] = fila_base.get(col)
                    dif[col + '_nuevo'] = fila_nuevo.get(col)
                diferencias.append(dif)
    
    if not diferencias:
        return pd.DataFrame()
    return pd.DataFrame(diferencias)

def obtener_prefijo_sheet(nombre, datos):
    xps_path = datos['rutas'].get('xps')
    if xps_path:
        parent_dir = os.path.dirname(xps_path)
        folder_name = os.path.basename(parent_dir)
        if folder_name and folder_name.strip() != "":
            return folder_name
    return nombre

def generar_nombre_hoja(prefijo, tipo, hojas_existentes):
    """
    Genera un nombre de hoja único y válido para Excel (máximo 31 caracteres) que:
    1. Conserve siempre el sufijo del tipo ('_RPS', '_SPS', '_XPS').
    2. Recorte el prefijo si excede el espacio disponible.
    3. Garantice que el nombre sea único en el libro (usando contador si hay colisión).
    """
    sufijo = f"_{tipo.strip()}" if not tipo.startswith('_') else tipo.strip()
    max_len_prefijo = 31 - len(sufijo)
    base_prefijo = prefijo[:max_len_prefijo].strip()
    candidato = f"{base_prefijo}{sufijo}"
    
    contador = 1
    while candidato in hojas_existentes:
        tag = f"_{contador}"
        max_pre = 31 - len(sufijo) - len(tag)
        candidato = f"{prefijo[:max_pre].strip()}{tag}{sufijo}"
        contador += 1
        
    hojas_existentes.add(candidato)
    return candidato

def exportar_excel_completo(paquetes, diferencias, xps_comparaciones=None, nombre_salida="comparacion_paquetes.xlsx"):
    """Exporta a Excel: resumen, datos originales de cada paquete (con encabezados descriptivos) y diferencias."""
    base, ext = os.path.splitext(nombre_salida)
    contador = 1
    nombre_final = nombre_salida
    while True:
        try:
            with pd.ExcelWriter(nombre_final, engine='openpyxl') as writer:
                hojas_existentes = set()
                
                # Hoja de resumen
                resumen = []
                for nombre, datos in paquetes.items():
                    prefijo = obtener_prefijo_sheet(nombre, datos)
                    rps_len = len(datos['rps'])
                    sps_len = len(datos['sps'])
                    xps_len = len(datos['xps'])
                    
                    status_parts = []
                    limite_export_res = 1040000
                    if rps_len > limite_export_res:
                        status_parts.append("RPS omitido (>1.04M)")
                    if sps_len > limite_export_res:
                        status_parts.append("SPS omitido (>1.04M)")
                    if xps_len > limite_export_res:
                        status_parts.append("XPS omitido (>1.04M)")
                        
                    export_status = "Completo" if not status_parts else "Parcial (" + ", ".join(status_parts) + ")"
                    
                    resumen.append({
                        'Paquete': prefijo,
                        'Carpeta': datos['carpeta'],
                        'Archivo RPS': datos['archivos']['rps'] or 'No encontrado',
                        'Archivo SPS': datos['archivos']['sps'] or 'No encontrado',
                        'Archivo XPS': datos['archivos']['xps'] or 'No encontrado',
                        'Total RPS': rps_len,
                        'Total SPS': sps_len,
                        'Total XPS': xps_len,
                        'Exportación Originales': export_status
                    })
                df_resumen = pd.DataFrame(resumen)
                sheet_resumen = 'Resumen Paquetes'
                df_resumen.to_excel(writer, sheet_name=sheet_resumen, index=False)
                aplicar_estilo_hoja(writer.sheets[sheet_resumen], '1F4E78')
                hojas_existentes.add(sheet_resumen)
                
                # Datos originales de cada paquete (por tipo) con encabezados descriptivos
                for nombre, datos in paquetes.items():
                    prefijo = obtener_prefijo_sheet(nombre, datos)
                    limite_export = 1040000
                    # RPS
                    if not datos['rps'].empty and len(datos['rps']) <= limite_export:
                        df = datos['rps'].copy()
                        df.rename(columns=ENCABEZADOS_RPS, inplace=True)
                        sheet_name = generar_nombre_hoja(prefijo, "RPS", hojas_existentes)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        aplicar_estilo_hoja(writer.sheets[sheet_name], '2F5597')
                    # SPS
                    if not datos['sps'].empty and len(datos['sps']) <= limite_export:
                        df = datos['sps'].copy()
                        df.rename(columns=ENCABEZADOS_SPS, inplace=True)
                        sheet_name = generar_nombre_hoja(prefijo, "SPS", hojas_existentes)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        aplicar_estilo_hoja(writer.sheets[sheet_name], '375623')
                    # XPS
                    if not datos['xps'].empty and len(datos['xps']) <= limite_export:
                        df = datos['xps'].copy()
                        df.rename(columns=ENCABEZADOS_XPS, inplace=True)
                        sheet_name = generar_nombre_hoja(prefijo, "XPS", hojas_existentes)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        aplicar_estilo_hoja(writer.sheets[sheet_name], '595959')
                
                # Hojas de diferencias con encabezados descriptivos
                for tipo, df in diferencias.items():
                    if tipo == 'RPS':
                        sheet_name = 'Diferencias RPS'
                        rename_map = {
                            'Paquete Base': 'Paquete Base',
                            'Paquete Nuevo': 'Paquete Nuevo',
                            'Tipo de Cambio': 'Tipo de Cambio',
                            'linea': 'Línea (2-17)',
                            'punto': 'Punto (18-25)',
                            'x_base': 'X Base (47-55)',
                            'x_nuevo': 'X Nuevo (47-55)',
                            'y_base': 'Y Base (56-65)',
                            'y_nuevo': 'Y Nuevo (56-65)',
                            'elevacion_base': 'Elevación Base (66-71)',
                            'elevacion_nuevo': 'Elevación Nuevo (66-71)'
                        }
                    elif tipo == 'SPS':
                        sheet_name = 'Diferencias SPS'
                        rename_map = {
                            'Paquete Base': 'Paquete Base',
                            'Paquete Nuevo': 'Paquete Nuevo',
                            'Tipo de Cambio': 'Tipo de Cambio',
                            'linea': 'Línea Fuente (2-17)',
                            'punto': 'Punto Fuente (18-25)',
                            'x_base': 'X Base (47-55)',
                            'x_nuevo': 'X Nuevo (47-55)',
                            'y_base': 'Y Base (56-65)',
                            'y_nuevo': 'Y Nuevo (56-65)',
                            'elevacion_base': 'Elevación Base (66-71)',
                            'elevacion_nuevo': 'Elevación Nuevo (66-71)'
                        }
                    else:
                        continue  # El XPS se maneja con la lógica estructural superior abajo
                    
                    if not df.empty:
                        df_renamed = df.rename(columns=rename_map)
                        # Reordenar columnas
                        cols = ['Paquete Base', 'Paquete Nuevo', 'Tipo de Cambio'] + [c for c in df_renamed.columns if c not in ['Paquete Base', 'Paquete Nuevo', 'Tipo de Cambio']]
                        df_renamed = df_renamed[cols]
                        df_renamed.to_excel(writer, sheet_name=sheet_name, index=False)
                        aplicar_estilo_hoja(writer.sheets[sheet_name], 'C00000', es_diferencias=True)
                
                # Escribir comparativas estructurales de XPS usando la lógica superior
                if xps_comparaciones:
                    escribir_hojas_comparativa_xps(writer, xps_comparaciones)
                
                break  # salir del while si se pudo escribir
        except PermissionError:
            # Si está abierto, cambiar nombre
            nombre_final = f"{base}_{contador}{ext}"
            contador += 1
            if contador > 100:
                raise RuntimeError("No se puede escribir el archivo Excel. Cierra el archivo abierto.")
    
    print(f" Excel generado: {nombre_final}")

def escribir_hojas_comparativa_xps(writer, xps_comparaciones):
    """
    Escribe los resultados de comparación de XPS (del comparador superior) al Excel.
    Genera dos hojas:
      - QC_XPS_Comp_Completa
      - QC_XPS_Diferencias
    """
    if not xps_comparaciones:
        return
        
    wb = writer.book
    
    # Colores y estilos inspirados en comparador_xps.py
    HEX = {
        "ok_bg":    "C8EDD6", "ok_fg":    "1B5E35",
        "diff_bg":  "FDECC8", "diff_fg":  "7B4D00",
        "only1_bg": "D0E8FB", "only1_fg": "0C3D6E",
        "only2_bg": "E8DFFB", "only2_fg": "3D1E87",
        "hdr_bg":   "2C3E50", "hdr_fg":   "FFFFFF",
        "alt_bg":   "F7F6F3", "white":    "FFFFFF",
        "border":   "BDB8AD",
        "miss_bg":  "FFB347", "miss_fg": "7A3500",
    }
    
    STATUS_STYLE = {
        "ok":    ("ok_bg",    "ok_fg"),
        "diff":  ("diff_bg",  "diff_fg"),
        "only1": ("only1_bg", "only1_fg"),
        "only2": ("only2_bg", "only2_fg"),
    }
    LINE_MISS_STYLE = ("miss_bg", "miss_fg")
    STATUS_LABEL = {
        "ok":    "Identico",
        "diff":  "Con diferencias",
        "only1": "Solo en Archivo 1",
        "only2": "Solo en Archivo 2",
    }
    
    _side = Side(style="thin", color=HEX["border"])
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)
    
    def _fill(h):  return PatternFill("solid", fgColor=h)
    def _font(h, bold=False, sz=10):
        return Font(color=h, bold=bold, size=sz, name="Calibri")
    def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    
    def apply_style(cell, bg=None, fg="1A1917", bold=False, align=None, sz=10):
        if bg: cell.fill = _fill(bg)
        cell.font      = _font(fg, bold=bold, sz=sz)
        cell.alignment = align or _left()
        cell.border    = _border
        
    def fmt_estacas(lst):
        if not lst: return "-"
        def _n(x):
            try: return float(x)
            except: return 0
        ini = min((e[0] for e in lst), key=_n)
        fin = max((e[1] for e in lst), key=_n)
        return f"{ini} -> {fin}" if len(lst) == 1 else f"{ini} -> {fin}  ({len(lst)} rangos)"

    HEADERS = ["COMPARATIVA", "DISPARO", "ESTADO DISPARO", "LÍNEA",
               "ESTACAS ARCHIVO 1", "ESTACAS ARCHIVO 2",
               "ESTADO LÍNEA", "DIFERENCIAS DE ESTACAS"]
               
    for sheet_name, only_diffs in [("QC_XPS_Comp_Completa", False), ("QC_XPS_Diferencias", True)]:
        ws = wb.create_sheet(sheet_name)
        ws.row_dimensions[1].height = 24
        ws.append(HEADERS)
        for col, h in enumerate(HEADERS, 1):
            apply_style(ws.cell(1, col), bg=HEX["hdr_bg"], fg=HEX["hdr_fg"],
                        bold=True, align=_center(), sz=10)
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = True
        
        alt = False
        row_count = 1
        
        for cmp in xps_comparaciones:
            par_name = cmp['par']
            result = cmp['resultado']
            
            for shot in result["results"]:
                s = shot["status"]
                if only_diffs and s == "ok":
                    continue
                bg_k, fg_k = STATUS_STYLE.get(s, ("alt_bg", "1A1917"))
                d_bg, d_fg = HEX[bg_k], HEX[fg_k]
                d_lbl = STATUS_LABEL.get(s, s)
                
                if s in ("only1", "only2"):
                    row_vals = [par_name, shot["disparo"], d_lbl, "-", "-", "-", "-", ""]
                    ws.append(row_vals)
                    r = ws.max_row
                    ws.row_dimensions[r].height = 18
                    for col, val in enumerate(row_vals, 1):
                        apply_style(ws.cell(r, col), bg=d_bg, fg=d_fg,
                                    bold=(col <= 3), align=_center() if col != 8 else _left())
                    row_count += 1
                    continue
                    
                for lr in shot["lineas"]:
                    ls = lr["status"]
                    if only_diffs and ls == "ok":
                        continue
                    if ls in ("only1", "only2"):
                        ls_bk, ls_fk = LINE_MISS_STYLE
                    else:
                        ls_bk, ls_fk = STATUS_STYLE.get(ls, ("alt_bg", "1A1917"))
                    l_bg, l_fg = HEX[ls_bk], HEX[ls_fk]
                    l_lbl = STATUS_LABEL.get(ls, ls)
                    
                    e1 = fmt_estacas(lr["estacas1"]) if lr["estacas1"] else "-"
                    e2 = fmt_estacas(lr["estacas2"]) if lr["estacas2"] else "-"
                    parts = []
                    for ed in lr["estaca_diffs"]:
                        a = "Arch.1" if ed["tipo"] == "solo_arch1" else "Arch.2"
                        parts.append(f"Solo {a}: {ed['ini']} -> {ed['fin']}")
                    diffs_txt = "  |  ".join(parts)
                    
                    alt = not alt
                    row_bg = HEX["alt_bg"] if alt else HEX["white"]
                    row_vals = [par_name, shot["disparo"], d_lbl, lr["linea"],
                                e1, e2, l_lbl, diffs_txt]
                                
                    ws.append(row_vals)
                    r = ws.max_row
                    ws.row_dimensions[r].height = 18
                    
                    if ls in ("only1", "only2"):
                        for col in range(1, 9):
                            apply_style(ws.cell(r, col), bg=l_bg, fg=l_fg,
                                        bold=(col <= 4), align=_center() if col != 8 else _left())
                    else:
                        apply_style(ws.cell(r, 1), bg=row_bg,                                   align=_center())
                        apply_style(ws.cell(r, 2), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                        apply_style(ws.cell(r, 3), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                        apply_style(ws.cell(r, 4), bg=l_bg,    fg=l_fg,        bold=True,         align=_center())
                        apply_style(ws.cell(r, 5), bg=row_bg,                                     align=_center())
                        apply_style(ws.cell(r, 6), bg=row_bg,                                     align=_center())
                        apply_style(ws.cell(r, 7), bg=l_bg,    fg=l_fg,        bold=(ls != "ok"), align=_center())
                        d_cell_bg = HEX["diff_bg"] if diffs_txt else row_bg
                        d_cell_fg = HEX["diff_fg"] if diffs_txt else "1A1917"
                        apply_style(ws.cell(r, 8), bg=d_cell_bg, fg=d_cell_fg,
                                    bold=bool(diffs_txt), align=_left())
                    row_count += 1
                    
        # Auto-ajustar anchos usando las primeras 200 filas
        filas_muestra = min(ws.max_row, 200)
        for col_idx in range(1, len(HEADERS) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for r_idx in range(1, filas_muestra + 1):
                val = ws.cell(row=r_idx, column=col_idx).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)