
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import csv
from collections import defaultdict


def read_file(path: str) -> list:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return [line.rstrip("\r\n") for line in f]


def parse_headers(lines: list) -> dict:
    h = {}
    for line in lines:
        if line.startswith("H"):
            code = line[:3].strip()
            h[code] = line[3:].strip()
    return h


def parse_structure(lines: list) -> dict:
    """
    Devuelve estructura jerarquizada:
      {
        disparo: {
          linea: [(estaca_ini, estaca_fin), ...]
        }
      }
    También devuelve set de disparos y dict de líneas por disparo.
    """
    structure = defaultdict(lambda: defaultdict(list))
    for line in lines:
        if not line.startswith("X"):
            continue
        fields = line.split()
        if len(fields) < 10:
            continue
        try:
            disparo     = fields[3]
            linea       = fields[7]
            estaca_ini  = fields[8]
            estaca_fin  = fields[9]
        except IndexError:
            continue
        structure[disparo][linea].append((estaca_ini, estaca_fin))
    return structure


def compare_structures(s1: dict, s2: dict, name1: str, name2: str) -> dict:
    """
    Compara las dos estructuras jerarquizadas.
    Devuelve un dict con todos los resultados organizados.
    """
    all_disparos = sorted(set(s1.keys()) | set(s2.keys()),
                          key=lambda x: float(x) if x.lstrip('-').replace('.','').isdigit() else x)

    results = []

    for disp in all_disparos:
        in1 = disp in s1
        in2 = disp in s2

        if not in1:
            results.append({
                "disparo": disp,
                "status": "only2",
                "lineas": [],
            })
            continue
        if not in2:
            results.append({
                "disparo": disp,
                "status": "only1",
                "lineas": [],
            })
            continue

        # Disparo existe en ambos — comparar líneas
        lineas1 = s1[disp]
        lineas2 = s2[disp]
        all_lineas = sorted(set(lineas1.keys()) | set(lineas2.keys()),
                            key=lambda x: float(x) if x.lstrip('-').replace('.','').isdigit() else x)

        linea_results = []
        disp_status = "ok"

        for lin in all_lineas:
            l_in1 = lin in lineas1
            l_in2 = lin in lineas2

            if not l_in1:
                linea_results.append({
                    "linea": lin, "status": "only2",
                    "estacas1": [], "estacas2": lineas2[lin],
                    "estaca_diffs": []
                })
                disp_status = "diff"
                continue
            if not l_in2:
                linea_results.append({
                    "linea": lin, "status": "only1",
                    "estacas1": lineas1[lin], "estacas2": [],
                    "estaca_diffs": []
                })
                disp_status = "diff"
                continue

            # Línea en ambos — comparar estacas
            set1 = set(lineas1[lin])
            set2 = set(lineas2[lin])

            estaca_diffs = []
            only_in_1 = sorted(set1 - set2, key=lambda x: (float(x[0]) if x[0].lstrip('-').replace('.','').isdigit() else x[0]))
            only_in_2 = sorted(set2 - set1, key=lambda x: (float(x[0]) if x[0].lstrip('-').replace('.','').isdigit() else x[0]))

            for e in only_in_1:
                estaca_diffs.append({"tipo": "solo_archivo 1", "ini": e[0], "fin": e[1]})
            for e in only_in_2:
                estaca_diffs.append({"tipo": "solo_archivo 2", "ini": e[0], "fin": e[1]})

            lin_status = "diff" if estaca_diffs else "ok"
            if lin_status == "diff":
                disp_status = "diff"

            linea_results.append({
                "linea": lin,
                "status": lin_status,
                "estacas1": sorted(lineas1[lin], key=lambda x: float(x[0]) if x[0].lstrip('-').replace('.','').isdigit() else x[0]),
                "estacas2": sorted(lineas2[lin], key=lambda x: float(x[0]) if x[0].lstrip('-').replace('.','').isdigit() else x[0]),
                "estaca_diffs": estaca_diffs,
            })

        results.append({
            "disparo": disp,
            "status": disp_status,
            "n_lineas1": len(lineas1),
            "n_lineas2": len(lineas2),
            "lineas": linea_results,
        })

    disparos_solo1   = [r for r in results if r["status"] == "only1"]
    disparos_solo2   = [r for r in results if r["status"] == "only2"]
    disparos_diff    = [r for r in results if r["status"] == "diff"]
    disparos_ok      = [r for r in results if r["status"] == "ok"]

    return {
        "results": results,
        "n_disparos1": len(s1),
        "n_disparos2": len(s2),
        "disparos_solo1": disparos_solo1,
        "disparos_solo2": disparos_solo2,
        "disparos_diff":  disparos_diff,
        "disparos_ok":    disparos_ok,
        "name1": name1,
        "name2": name2,
    }


def run_comparison(path1: str, path2: str) -> dict:
    lines1 = read_file(path1)
    lines2 = read_file(path2)
    h1, h2 = parse_headers(lines1), parse_headers(lines2)
    s1 = parse_structure(lines1)
    s2 = parse_structure(lines2)
    cmp = compare_structures(s1, s2,
                             os.path.basename(path1),
                             os.path.basename(path2))
    cmp["h_diffs"] = []
    for k in sorted(set(h1) | set(h2)):
        v1 = h1.get(k, "<ausente>")
        v2 = h2.get(k, "<ausente>")
        if v1 != v2:
            cmp["h_diffs"].append({"code": k, "v1": v1, "v2": v2})
    return cmp


def export_txt(result: dict, path: str):
    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    with open(path, "w", encoding="utf-8") as f:
        f.write("REPORTE DE DIFERENCIAS SPS/XPS\n")
        f.write(f"Archivo 1 : {result['name1']}  ({n1} disparos)\n")
        f.write(f"Archivo 2 : {result['name2']}  ({n2} disparos)\n")
        f.write("=" * 72 + "\n\n")

        if n1 != n2:
            f.write(f"⚠ DIFERENCIA EN DISPAROS: Archivo 1={n1}, Arch2={n2}\n\n")

        if result["h_diffs"]:
            f.write("CABECERAS DISTINTAS:\n")
            for d in result["h_diffs"]:
                f.write(f"  [{d['code']}]  Archivo 1: {d['v1']}  →  Archivo 2: {d['v2']}\n")
            f.write("\n")

        if result["disparos_solo1"]:
            f.write("DISPAROS SOLO EN ARCHIVO 1:\n")
            for r in result["disparos_solo1"]:
                f.write(f"  Disparo {r['disparo']}\n")
            f.write("\n")

        if result["disparos_solo2"]:
            f.write("DISPAROS SOLO EN ARCHIVO 2:\n")
            for r in result["disparos_solo2"]:
                f.write(f"  Disparo {r['disparo']}\n")
            f.write("\n")

        for r in result["results"]:
            if r["status"] == "ok":
                continue
            if r["status"] in ("only1", "only2"):
                arch = "Archivo 1" if r["status"] == "only1" else "Archivo 2"
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en {arch}\n\n")
                continue

            f.write(f"DISPARO {r['disparo']}\n")
            if r.get("n_lineas1") != r.get("n_lineas2"):
                f.write(f"  ⚠ Líneas: Arch1={r['n_lineas1']}, Arch2={r['n_lineas2']}\n")

            for lr in r["lineas"]:
                if lr["status"] == "ok":
                    continue
                if lr["status"] == "only1":
                    f.write(f"  Línea {lr['linea']}  ⚠ Solo en Archivo 1\n")
                    for e in lr["estacas1"]:
                        f.write(f"    Estacas {e[0]} → {e[1]}\n")
                elif lr["status"] == "only2":
                    f.write(f"  Línea {lr['linea']}  ⚠ Solo en Archivo 2\n")
                    for e in lr["estacas2"]:
                        f.write(f"    Estacas {e[0]} → {e[1]}\n")
                else:
                    f.write(f"  Línea {lr['linea']}  ({len(lr['estaca_diffs'])} diferencia(s) en estacas)\n")
                    for ed in lr["estaca_diffs"]:
                        arch = "Archivo 1" if ed["tipo"] == "solo_archivo 1" else "Archivo 2"
                        f.write(f"    Solo en {arch}: estaca {ed['ini']} → {ed['fin']}\n")
            f.write("\n")


def export_csv(result: dict, path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["disparo", "linea", "tipo_diferencia",
                    "archivo", "estaca_ini", "estaca_fin"])
        for r in result["results"]:
            if r["status"] == "ok":
                continue
            if r["status"] == "only1":
                w.writerow([r["disparo"], "—", "disparo_faltante",
                             "Archivo 2", "—", "—"])
                continue
            if r["status"] == "only2":
                w.writerow([r["disparo"], "—", "disparo_faltante",
                             "Archivo 1", "—", "—"])
                continue
            for lr in r["lineas"]:
                if lr["status"] == "ok":
                    continue
                if lr["status"] == "only1":
                    for e in lr["estacas1"]:
                        w.writerow([r["disparo"], lr["linea"], "linea_faltante",
                                    "Archivo 2", e[0], e[1]])
                elif lr["status"] == "only2":
                    for e in lr["estacas2"]:
                        w.writerow([r["disparo"], lr["linea"], "linea_faltante",
                                    "Archivo 1", e[0], e[1]])
                else:
                    for ed in lr["estaca_diffs"]:
                        arch = "Archivo 1" if ed["tipo"] == "solo_arch1" else "Archivo 2"
                        w.writerow([r["disparo"], lr["linea"], "estaca_faltante",
                                    arch, ed["ini"], ed["fin"]])


def export_comparativa(result: dict, path: str):
    """
    Exporta la comparativa completa: todos los disparos, todas las líneas
    y todas las estacas — tanto coincidentes como diferentes.
    """
    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    W = 76

    def sep(ch="─"): return ch * W

    with open(path, "w", encoding="utf-8") as f:
        f.write("COMPARATIVA COMPLETA SPS/XPS\n")
        f.write(f"Archivo 1 : {result['name1']}  ({n1} disparos)\n")
        f.write(f"Archivo 2 : {result['name2']}  ({n2} disparos)\n")
        f.write(sep("═") + "\n\n")

        # Resumen ejecutivo
        ndiff  = len(result["disparos_diff"])
        nmiss1 = len(result["disparos_solo1"])
        nmiss2 = len(result["disparos_solo2"])
        nok    = len(result["disparos_ok"])
        f.write("RESUMEN\n")
        f.write(f"  Disparos totales comparados : {len(result['results'])}\n")
        f.write(f"  Idénticos                   : {nok}\n")
        f.write(f"  Con diferencias de líneas/estacas : {ndiff}\n")
        f.write(f"  Solo en Archivo 1           : {nmiss1}\n")
        f.write(f"  Solo en Archivo 2           : {nmiss2}\n")
        if n1 != n2:
            diff = abs(n1 - n2)
            which = "Archivo 1" if n1 > n2 else "Archivo 2"
            f.write(f"  ⚠ {which} tiene {diff} disparo(s) extra\n")
        f.write("\n")

        if result["h_diffs"]:
            f.write("DIFERENCIAS EN CABECERAS (H)\n")
            f.write(sep() + "\n")
            for d in result["h_diffs"]:
                f.write(f"  [{d['code']}]\n")
                f.write(f"    Archivo 1: {d['v1']}\n")
                f.write(f"    Archivo 2: {d['v2']}\n")
            f.write("\n")

        # Detalle por disparo
        for r in result["results"]:
            f.write(sep("═") + "\n")
            if r["status"] == "only1":
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en Archivo 1\n\n")
                continue
            if r["status"] == "only2":
                f.write(f"DISPARO {r['disparo']}  ⚠ Solo en Archivo 2\n\n")
                continue

            status_lbl = "✔ Idéntico" if r["status"] == "ok" else "⚠ Con diferencias"
            f.write(f"DISPARO {r['disparo']}  [{status_lbl}]  "
                    f"(Arch1: {r.get('n_lineas1','?')} líneas  |  "
                    f"Arch2: {r.get('n_lineas2','?')} líneas)\n")
            f.write(sep() + "\n")

            # Encabezado de tabla
            f.write(f"  {'Línea':<10} {'Estado':<22} {'Estacas Archivo 1':<22} "
                    f"{'Estacas Archivo 2':<22} Diferencias\n")
            f.write(f"  {'-' * (W-2)}\n")

            for lr in r["lineas"]:
                s = lr["status"]

                def fmt(lst):
                    if not lst: return "—"
                    if len(lst) == 1: return f"{lst[0][0]}→{lst[0][1]}"
                    def _n(x):
                        return float(x) if x.lstrip('-').replace('.','').isdigit() else 0
                    ini = min((e[0] for e in lst), key=_n)
                    fin = max((e[1] for e in lst), key=_n)
                    return f"{ini}→{fin} ({len(lst)})"

                if s == "only1":
                    estado  = "Solo en Archivo 1"
                    e1_txt  = fmt(lr["estacas1"])
                    e2_txt  = "—"
                    diff_tx = ""
                elif s == "only2":
                    estado  = "Solo en Archivo 2"
                    e1_txt  = "—"
                    e2_txt  = fmt(lr["estacas2"])
                    diff_tx = ""
                else:
                    estado  = "✔ ok" if s == "ok" else f"⚠ {len(lr['estaca_diffs'])} dif."
                    e1_txt  = fmt(lr["estacas1"])
                    e2_txt  = fmt(lr["estacas2"])
                    parts   = []
                    for ed in lr["estaca_diffs"]:
                        a = "Arch1" if ed["tipo"] == "solo_arch1" else "Arch2"
                        parts.append(f"{a}:{ed['ini']}→{ed['fin']}")
                    diff_tx = " | ".join(parts)

                f.write(f"  {lr['linea']:<10} {estado:<22} {e1_txt:<22} "
                        f"{e2_txt:<22} {diff_tx}\n")

            f.write("\n")


def export_xlsx(result: dict, path: str):
    """
    Exporta la comparativa completa a Excel (.xlsx) con formato visual.
    3 hojas: Resumen, Comparativa completa, Solo diferencias.
    Requiere: openpyxl  (pip install openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "Se necesita openpyxl.\nInstálalo con:  pip install openpyxl")

    HEX = {
        "ok_bg":    "C8EDD6", "ok_fg":    "1B5E35",
        "diff_bg":  "FDECC8", "diff_fg":  "7B4D00",
        "only1_bg": "D0E8FB", "only1_fg": "0C3D6E",
        "only2_bg": "E8DFFB", "only2_fg": "3D1E87",
        "hdr_bg":   "2C3E50", "hdr_fg":   "FFFFFF",
        "sub_bg":   "546E7A", "sub_fg":   "FFFFFF",
        "alt":      "F5F4F1", "white":    "FFFFFF",
        "border":   "BDB8AD",
        "miss_bg":  "FFB347", "miss_fg": "7A3500",
    }

    STATUS_STYLE = {
        "ok":    ("ok_bg",    "ok_fg"),
        "diff":  ("diff_bg",  "diff_fg"),
        "only1": ("only1_bg", "only1_fg"),
        "only2": ("only2_bg", "only2_fg"),
    }
    LINE_MISS_STYLE = ("miss_bg", "miss_fg")   # naranja para líneas faltantes
    STATUS_LABEL = {
        "ok":    "Identico",
        "diff":  "Con diferencias",
        "only1": "Solo en Archivo 1",
        "only2": "Solo en Archivo 2",
    }

    def _fill(h):
        return PatternFill("solid", fgColor=h)

    def _font(h, bold=False, sz=10):
        return Font(color=h, bold=bold, size=sz, name="Calibri")

    def _border():
        s = Side(style="thin", color=HEX["border"])
        return Border(left=s, right=s, top=s, bottom=s)

    def _ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _lft():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style(cell, bg=None, fg="1A1917", bold=False, sz=10, align=None, brd=True):
        if bg:
            cell.fill = _fill(bg)
        cell.font  = _font(fg, bold=bold, size=sz)
        cell.alignment = align or _lft()
        if brd:
            cell.border = _border()

    def fmt_estacas(lst):
        if not lst:
            return "-"
        def _n(x):
            try: return float(x)
            except: return 0
        ini = min((e[0] for e in lst), key=_n)
        fin = max((e[1] for e in lst), key=_n)
        return f"{ini} - {fin}" if len(lst) == 1 else f"{ini} - {fin} ({len(lst)} rangos)"

    wb = Workbook()

    # ── HOJA 1: RESUMEN ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20 # pyright: ignore[reportOptionalMemberAccess]
    ws.column_dimensions["C"].width = 26

    # Título
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "COMPARATIVA SPS / XPS"
    style(c, bg=HEX["hdr_bg"], fg=HEX["hdr_fg"], bold=True, sz=14, align=_ctr())
    ws.row_dimensions[1].height = 34

    # Archivos
    for i, (lbl, val) in enumerate([
        ("Archivo 1:", result["name1"]),
        ("Archivo 2:", result["name2"]),
    ], 2):
        ws.merge_cells(f"A{i}:C{i}")
        c = ws.cell(i, 1, f"  {lbl}   {val}")
        style(c, bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True)
        ws.row_dimensions[i].height = 18

    ws.append([])  # fila vacía

    # Cabecera métricas
    ws.append(["INDICADOR", "ARCHIVO 1", "ARCHIVO 2"])
    r = ws.max_row
    ws.row_dimensions[r].height = 20
    for col in range(1, 4):
        c = ws.cell(r, col)
        style(c, bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True, align=_ctr())

    n1 = result["n_disparos1"]
    n2 = result["n_disparos2"]
    ndiff  = len(result["disparos_diff"])
    nmiss1 = len(result["disparos_solo1"])
    nmiss2 = len(result["disparos_solo2"])
    nok    = len(result["disparos_ok"])

    filas_met = [
        ("Total de disparos",               n1,    n2),
        ("Disparos identicos",              nok,   nok),
        ("Disparos con difs. lineas/estacas", ndiff, ndiff),
        ("Disparos faltantes",              nmiss1, nmiss2),
    ]
    for i, (label, v1, v2) in enumerate(filas_met):
        ws.append([label, v1, v2])
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        bg = HEX["alt"] if i % 2 else HEX["white"]
        for col in range(1, 4):
            c = ws.cell(r, col)
            style(c, bg=bg, align=_ctr() if col > 1 else _lft())
        # Colorear si hay problemas
        if label.startswith("Disparos con") and ndiff > 0:
            for col in (2, 3):
                c = ws.cell(r, col)
                style(c, bg=HEX["diff_bg"], fg=HEX["diff_fg"], bold=True, align=_ctr())
        if label.startswith("Disparos faltantes"):
            if nmiss1 > 0:
                c = ws.cell(r, 2)
                style(c, bg=HEX["only1_bg"], fg=HEX["only1_fg"], bold=True, align=_ctr())
            if nmiss2 > 0:
                c = ws.cell(r, 3)
                style(c, bg=HEX["only2_bg"], fg=HEX["only2_fg"], bold=True, align=_ctr())

    # ── FUNCIÓN HELPER para hojas de datos ───────────────────────────────
    HEADERS = ["DISPARO", "ESTADO DISPARO", "LINEA",
               "ESTACAS ARCHIVO 1", "ESTACAS ARCHIVO 2",
               "ESTADO LINEA", "DIFERENCIAS DE ESTACAS"]
    COL_W   = [12, 18, 12, 22, 22, 18, 36]

    def build_sheet(ws_dest, shots):
        for col, w in enumerate(COL_W, 1):
            ws_dest.column_dimensions[get_column_letter(col)].width = w
        ws_dest.freeze_panes = "A2"

        # Cabecera
        ws_dest.append(HEADERS)
        ws_dest.row_dimensions[1].height = 22
        for col in range(1, len(HEADERS)+1):
            c = ws_dest.cell(1, col)
            style(c, bg=HEX["hdr_bg"], fg=HEX["hdr_fg"], bold=True, align=_ctr())

        alt = False
        for shot in shots:
            s      = shot["status"]
            bg_d   = HEX[STATUS_STYLE[s][0]]
            fg_d   = HEX[STATUS_STYLE[s][1]]
            lbl_d  = STATUS_LABEL[s]

            if s in ("only1", "only2"):
                ws_dest.append([shot["disparo"], lbl_d, "-", "-", "-", "-", ""])
                r = ws_dest.max_row
                ws_dest.row_dimensions[r].height = 16
                for col in range(1, 8):
                    c = ws_dest.cell(r, col)
                    style(c, bg=bg_d, fg=fg_d, bold=(col <= 2), align=_ctr())
                continue

            for lr in shot["lineas"]:
                ls     = lr["status"]
                bg_l   = HEX[STATUS_STYLE[ls][0]]
                fg_l   = HEX[STATUS_STYLE[ls][1]]
                lbl_l  = STATUS_LABEL[ls]
                e1     = fmt_estacas(lr["estacas1"])
                e2     = fmt_estacas(lr["estacas2"])
                diffs  = [
                    f"{'Arch.1' if ed['tipo']=='solo_arch1' else 'Arch.2'}: {ed['ini']}-{ed['fin']}"
                    for ed in lr["estaca_diffs"]
                ]
                diff_txt = "  |  ".join(diffs)
                alt = not alt
                row_bg = HEX["alt"] if alt else HEX["white"]

                ws_dest.append([shot["disparo"], lbl_d,
                                lr["linea"], e1, e2, lbl_l, diff_txt])
                r = ws_dest.max_row
                ws_dest.row_dimensions[r].height = 16

                # DISPARO
                c = ws_dest.cell(r, 1)
                style(c, bg=bg_d, fg=fg_d, bold=True, align=_ctr())
                # ESTADO DISPARO
                c = ws_dest.cell(r, 2)
                style(c, bg=bg_d, fg=fg_d, bold=True, align=_ctr())
                # LINEA
                c = ws_dest.cell(r, 3)
                style(c, bg=bg_l, fg=fg_l, bold=True, align=_ctr())
                # ESTACAS 1 y 2
                for col in (4, 5):
                    c = ws_dest.cell(r, col)
                    style(c, bg=row_bg, align=_ctr())
                # ESTADO LINEA
                c = ws_dest.cell(r, 6)
                style(c, bg=bg_l, fg=fg_l, bold=(ls != "ok"), align=_ctr())
                # DIFERENCIAS
                c = ws_dest.cell(r, 7)
                if diff_txt:
                    style(c, bg=HEX["diff_bg"], fg=HEX["diff_fg"], bold=True, align=_lft())
                else:
                    style(c, bg=row_bg, align=_lft())

    # ── HOJA 2: COMPARATIVA COMPLETA ─────────────────────────────────────
    ws2 = wb.create_sheet("Comparativa completa")
    build_sheet(ws2, result["results"])

    # ── HOJA 3: SOLO DIFERENCIAS ──────────────────────────────────────────
    ws3 = wb.create_sheet("Solo diferencias")
    shots_diff = [s for s in result["results"] if s["status"] != "ok"]
    if shots_diff:
        build_sheet(ws3, shots_diff)
    else:
        ws3["A1"].value = "No se encontraron diferencias."
        style(ws3["A1"], bg=HEX["ok_bg"], fg=HEX["ok_fg"], bold=True)

    wb.save(path)



def export_xlsx(result: dict, path: str):
    """
    Exporta la comparativa completa a Excel (.xlsx) con formato visual.
    3 hojas: Resumen · Comparativa completa · Solo diferencias
    Requiere: openpyxl  (pip install openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "Se necesita openpyxl para exportar a Excel.\n"
            "Instálalo con:  pip install openpyxl")

    HEX = {
        "ok_bg":    "C8EDD6", "ok_fg":    "1B5E35",
        "diff_bg":  "FDECC8", "diff_fg":  "7B4D00",
        "only1_bg": "D0E8FB", "only1_fg": "0C3D6E",
        "only2_bg": "E8DFFB", "only2_fg": "3D1E87",
        "hdr_bg":   "2C3E50", "hdr_fg":   "FFFFFF",
        "sub_bg":   "546E7A", "sub_fg":   "FFFFFF",
        "alt_bg":   "F7F6F3", "white":    "FFFFFF",
        "border":   "BDB8AD",
        "miss_bg":  "FFB347", "miss_fg": "7A3500",
    }

    STATUS_STYLE = {
        "ok":    ("ok_bg",    "ok_fg"),
        "diff":  ("diff_bg",  "diff_fg"),
        "only1": ("only1_bg", "only1_fg"),
        "only2": ("only2_bg", "only2_fg"),
    }
    LINE_MISS_STYLE = ("miss_bg", "miss_fg")   # naranja para líneas faltantes
    STATUS_LABEL = {
        "ok":    "✔ Idéntico",
        "diff":  "⚠ Con diferencias",
        "only1": "✖ Solo en Archivo 1",
        "only2": "✖ Solo en Archivo 2",
    }

    _side = Side(style="thin", color=HEX["border"])
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    def _fill(h):  return PatternFill("solid", fgColor=h)
    def _font(h, bold=False, sz=10):
        return Font(color=h, bold=bold, size=sz, name="Calibri")
    def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def apply(cell, bg=None, fg="1A1917", bold=False, align=None, sz=10):
        if bg: cell.fill = _fill(bg)
        cell.font      = _font(fg, bold=bold, sz=sz)
        cell.alignment = align or _left()
        cell.border    = _border

    def fmt_estacas(lst):
        if not lst: return "—"
        def _n(x):
            try: return float(x)
            except: return 0
        ini = min((e[0] for e in lst), key=_n)
        fin = max((e[1] for e in lst), key=_n)
        return f"{ini} → {fin}" if len(lst) == 1 else f"{ini} → {fin}  ({len(lst)} rangos)"

    HEADERS = ["DISPARO", "ESTADO DISPARO", "LÍNEA",
               "ESTACAS ARCHIVO 1", "ESTACAS ARCHIVO 2",
               "ESTADO LÍNEA", "DIFERENCIAS DE ESTACAS"]

    # Rastreo de ancho máximo por columna para auto-ajuste
    _col_max = {}   # sheet_title -> {col_idx: max_len}

    def _track(ws_title, col, value):
        """Registra la longitud de celda para calcular el ancho óptimo."""
        length = len(str(value)) if value else 0
        if ws_title not in _col_max:
            _col_max[ws_title] = {}
        if col not in _col_max[ws_title] or length > _col_max[ws_title][col]:
            _col_max[ws_title][col] = length

    def auto_fit(ws):
        """Aplica el ancho óptimo a cada columna según el contenido registrado."""
        col_lens = _col_max.get(ws.title, {})
        for col, max_len in col_lens.items():
            # +4 de margen; mínimo 10, máximo 60
            width = min(max(max_len + 4, 10), 60)
            ws.column_dimensions[get_column_letter(col)].width = width

    def setup_sheet(ws, title_suffix=""):
        ws.row_dimensions[1].height = 24
        ws.append(HEADERS)
        for col, h in enumerate(HEADERS, 1):
            apply(ws.cell(1, col), bg=HEX["hdr_bg"], fg=HEX["hdr_fg"],
                  bold=True, align=_center(), sz=10)
            _track(ws.title, col, h)
        ws.freeze_panes = "A2"

    def write_shot_rows(ws, shot_list, only_diffs=False):
        alt = False
        for shot in shot_list:
            s = shot["status"]
            if only_diffs and s == "ok":
                continue
            bg_k, fg_k = STATUS_STYLE.get(s, ("alt_bg", "1A1917"))
            d_bg, d_fg = HEX[bg_k], HEX[fg_k]
            d_lbl = STATUS_LABEL.get(s, s)

            if s in ("only1", "only2"):
                row_vals = [shot["disparo"], d_lbl, "—", "—", "—", "—", ""]
                ws.append(row_vals)
                r = ws.max_row
                ws.row_dimensions[r].height = 18
                for col, val in enumerate(row_vals, 1):
                    apply(ws.cell(r, col), bg=d_bg, fg=d_fg,
                          bold=(col <= 2), align=_center())
                    _track(ws.title, col, val)
                continue

            for lr in shot["lineas"]:
                ls = lr["status"]
                if only_diffs and ls == "ok":
                    continue
                # Líneas faltantes en un archivo → naranja
                if ls in ("only1", "only2"):
                    ls_bk, ls_fk = LINE_MISS_STYLE
                else:
                    ls_bk, ls_fk = STATUS_STYLE.get(ls, ("alt_bg", "1A1917"))
                l_bg, l_fg = HEX[ls_bk], HEX[ls_fk]
                l_lbl = STATUS_LABEL.get(ls, ls)
                # Mostrar estacas del archivo que sí las tiene
                e1 = fmt_estacas(lr["estacas1"]) if lr["estacas1"] else "—"
                e2 = fmt_estacas(lr["estacas2"]) if lr["estacas2"] else "—"
                parts = []
                for ed in lr["estaca_diffs"]:
                    a = "Arch.1" if ed["tipo"] == "solo_arch1" else "Arch.2"
                    parts.append(f"Solo {a}: {ed['ini']}→{ed['fin']}")
                diffs_txt = "  |  ".join(parts)

                alt = not alt
                row_bg = HEX["alt_bg"] if alt else HEX["white"]
                row_vals = [shot["disparo"], d_lbl, lr["linea"],
                            e1, e2, l_lbl, diffs_txt]

                ws.append(row_vals)
                r = ws.max_row
                ws.row_dimensions[r].height = 18

                for col, val in enumerate(row_vals, 1):
                    _track(ws.title, col, val)

                # Si la línea es faltante, pintar toda la fila en naranja
                if ls in ("only1", "only2"):
                    for col in range(1, 8):
                        apply(ws.cell(r, col), bg=l_bg, fg=l_fg,
                              bold=(col <= 3), align=_center())
                    # Columna 7 sin contenido de diferencias
                    apply(ws.cell(r, 7), bg=l_bg, fg=l_fg, align=_left())
                else:
                    apply(ws.cell(r, 1), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 2), bg=d_bg,    fg=d_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 3), bg=l_bg,    fg=l_fg,        bold=True,         align=_center())
                    apply(ws.cell(r, 4), bg=row_bg,                                     align=_center())
                    apply(ws.cell(r, 5), bg=row_bg,                                     align=_center())
                    apply(ws.cell(r, 6), bg=l_bg,    fg=l_fg,        bold=(ls != "ok"), align=_center())
                    d_cell_bg = HEX["diff_bg"] if diffs_txt else row_bg
                    d_cell_fg = HEX["diff_fg"] if diffs_txt else "1A1917"
                    apply(ws.cell(r, 7), bg=d_cell_bg, fg=d_cell_fg,
                          bold=bool(diffs_txt), align=_left())

    wb = Workbook()

    # ── Hoja 1: Resumen ──────────────────────────────────────────────────
    ws_r = wb.active
    ws_r.title = "Resumen"
    ws_r.column_dimensions["A"].width = 38
    ws_r.column_dimensions["B"].width = 18
    ws_r.column_dimensions["C"].width = 28

    ws_r.merge_cells("A1:C1")
    apply(ws_r["A1"], bg=HEX["hdr_bg"], fg=HEX["hdr_fg"],
          bold=True, align=_center(), sz=14)
    ws_r["A1"].value = "COMPARATIVA SPS / XPS"
    ws_r.row_dimensions[1].height = 36

    for i, (lbl, val) in enumerate([
        ("Archivo 1", result["name1"]),
        ("Archivo 2", result["name2"]),
    ], start=2):
        ws_r.merge_cells(f"A{i}:C{i}")
        c = ws_r.cell(i, 1, f"  {lbl}:   {val}")
        apply(c, bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True)
        ws_r.row_dimensions[i].height = 20
        _track(ws_r.title, 1, f"  {lbl}:   {val}")

    ws_r.append([])
    ws_r.append(["INDICADOR", "VALOR", "NOTA"])
    for col in range(1, 4):
        apply(ws_r.cell(ws_r.max_row, col),
              bg=HEX["sub_bg"], fg=HEX["sub_fg"], bold=True, align=_center())
    ws_r.row_dimensions[ws_r.max_row].height = 20

    n1, n2 = result["n_disparos1"], result["n_disparos2"]
    ndiff  = len(result["disparos_diff"])
    nmiss1 = len(result["disparos_solo1"])
    nmiss2 = len(result["disparos_solo2"])
    nok    = len(result["disparos_ok"])

    rows_sum = [
        ("Disparos en Archivo 1",             n1,     ""),
        ("Disparos en Archivo 2",             n2,
         "⚠ Cantidades distintas" if n1 != n2 else "✔ Coinciden"),
        ("Disparos idénticos",                nok,    ""),
        ("Disparos con difs. líneas/estacas", ndiff,  ""),
        ("Solo en Archivo 1",                 nmiss1, ""),
        ("Solo en Archivo 2",                 nmiss2, ""),
    ]
    for i, (lbl, val, nota) in enumerate(rows_sum):
        ws_r.append([lbl, val, nota])
        r = ws_r.max_row
        row_bg = HEX["alt_bg"] if i % 2 == 0 else HEX["white"]
        for col, v in enumerate([lbl, val, nota], 1):
            apply(ws_r.cell(r, col), bg=row_bg)
            _track(ws_r.title, col, v)
        # Resaltar valores problemáticos
        if "difs" in lbl and val > 0:
            apply(ws_r.cell(r, 2), bg=HEX["diff_bg"], fg=HEX["diff_fg"], bold=True, align=_center())
        elif "Solo en" in lbl and val > 0:
            apply(ws_r.cell(r, 2), bg=HEX["only1_bg"], fg=HEX["only1_fg"], bold=True, align=_center())
        elif lbl == "Disparos idénticos" and val == n1 == n2:
            apply(ws_r.cell(r, 2), bg=HEX["ok_bg"], fg=HEX["ok_fg"], bold=True, align=_center())
        ws_r.row_dimensions[r].height = 18

    # ── Hoja 2: Comparativa completa ─────────────────────────────────────
    ws_full = wb.create_sheet("Comparativa completa")
    setup_sheet(ws_full)
    write_shot_rows(ws_full, result["results"], only_diffs=False)
    auto_fit(ws_full)

    # ── Hoja 3: Solo diferencias ─────────────────────────────────────────
    ws_diff = wb.create_sheet("Solo diferencias")
    setup_sheet(ws_diff)
    write_shot_rows(ws_diff, result["results"], only_diffs=True)
    auto_fit(ws_diff)

    # Auto-fit hoja resumen
    auto_fit(ws_r)

    wb.save(path)


# Interfaz

COLORS = {
    "bg":         "#F2F1EE",
    "surface":    "#FFFFFF",
    "border":     "#D8D4CC",
    "text":       "#1A1917",
    "muted":      "#6B6860",
    "accent":     "#1A1917",
    # estados — fondo sólido para filas
    "ok_row":     "#139244",   # verde suave
    "ok_txt":     "#032B0F",
    "diff_row":   "#D6D63B",   # amarillo suave
    "diff_txt":   "#505241",
    "only1_row":  "#57A3E0",   # azul suave
    "only1_txt":  "#09032C",
    "only2_row":  "#8352EC",   # violeta suave
    "only2_txt":  "#031F30",
    # cabeceras de columna 
    "col_hdr":    "#E8E5DF",
    "col_hdr_txt":"#444240",
}

FONT      = ("Segoe UI", 10)
FONT_BOLD = ("Segoe UI", 10, "bold")
FONT_SM   = ("Segoe UI", 9)
FONT_H1   = ("Segoe UI", 14, "bold")
FONT_MONO = ("Consolas", 9)

# Anchos de columna (proporciones)
COL_DISP  = 140   # Disparo
COL_ARCH  = 1     # Archivo 1 y 2 (expand igual)

STATUS_META = {
    "ok":    ("✔ Idéntico",       "ok_row",    "ok_txt"),
    "diff":  ("⚠ Con diferencias","diff_row",  "diff_txt"),
    "only1": ("✖ Solo en Archivo 1","only1_row", "only1_txt"),
    "only2": ("✖ Solo en Archivo 2","only2_row", "only2_txt"),
}


# ─── Selector de archivo ──────────────────────────────────────────────────────
class FileSelector(tk.Frame):
    def __init__(self, parent, label, **kw):
        super().__init__(parent, bg=COLORS["surface"],
                         highlightthickness=1,
                         highlightbackground=COLORS["border"], **kw)
        self._path = tk.StringVar()
        self._label = label
        self._build()

    def _build(self):
        tk.Label(self, text=self._label, font=FONT_BOLD,
                 bg=COLORS["surface"], fg=COLORS["text"]).pack(pady=(12, 3))
        self._icon = tk.Label(self, text="📂", font=("Segoe UI", 26),
                              bg=COLORS["surface"])
        self._icon.pack()
        self._info = tk.Label(self,
                              text="Haz clic para seleccionar .xps / .sps",
                              font=FONT_SM, bg=COLORS["surface"],
                              fg=COLORS["muted"], wraplength=220)
        self._info.pack(pady=(3, 8))
        tk.Button(self, text="Seleccionar archivo de disparo",
                  font=FONT_SM, cursor="hand2",
                  bg=COLORS["bg"], fg=COLORS["text"],
                  relief="flat", bd=1,
                  activebackground=COLORS["border"],
                  command=self._pick).pack(pady=(0, 12))

    def _pick(self):
        p = filedialog.askopenfilename(
            filetypes=[("SPS/XPS", "*.xps *.sps *.txt"), ("Todos", "*.*")])
        if not p:
            return
        self._path.set(p)
        name = os.path.basename(p)
        try:
            with open(p, "r", errors="replace") as f:
                lcount = sum(1 for _ in f)
            xcount = sum(1 for ln in open(p, "r", errors="replace")
                         if ln.startswith("X"))
        except Exception:
            lcount = xcount = 0
        self._info.configure(
            text=f"{name}\n{lcount:,} líneas  ·  {xcount:,} registros X",
            fg=COLORS["ok_txt"])
        self._icon.configure(text="✅")

    def get(self):
        return self._path.get()


# ─── Tarjeta de métrica ───────────────────────────────────────────────────────
class MetricCard(tk.Frame):
    def __init__(self, parent, label, **kw):
        super().__init__(parent, bg=COLORS["surface"],
                         highlightthickness=1,
                         highlightbackground=COLORS["border"], **kw)
        tk.Label(self, text=label, font=("Segoe UI", 8),
                 bg=COLORS["surface"], fg=COLORS["muted"]).pack(pady=(10, 2), padx=10)
        self._val = tk.Label(self, text="—", font=("Segoe UI", 18, "bold"),
                             bg=COLORS["surface"], fg=COLORS["text"])
        self._val.pack(pady=(0, 10))

    def set(self, value, color=None):
        self._val.configure(text=str(value), fg=color or COLORS["text"])


# ─── Fila de disparo (3 columnas) ────────────────────────────────────────────
class DisparoRow(tk.Frame):
    """
    Tres celdas horizontales:
      [DISPARO #]  |  [contenido Archivo 1]  |  [contenido Archivo 2]
    Al hacer clic se expande el detalle de líneas.
    """

    def __init__(self, parent, r: dict, **kw):
        super().__init__(parent, bg=COLORS["bg"], **kw)
        self._r = r
        self._open = False
        self._detail = None
        self._build_header()

    def _meta(self):
        return STATUS_META.get(self._r["status"], STATUS_META["diff"])

    def _build_header(self):
        r = self._r
        lbl, bg_key, fg_key = self._meta()
        bg = COLORS[bg_key]
        fg = COLORS[fg_key]

        self._hdr = tk.Frame(self, bg=bg, cursor="hand2")
        self._hdr.pack(fill="x")
        self._hdr.bind("<Button-1>", self._toggle)

        # ── Columna DISPARO (izquierda) ──
        disp_cell = tk.Frame(self._hdr, bg=bg, width=COL_DISP)
        disp_cell.pack(side="left", fill="y")
        disp_cell.pack_propagate(False)
        tk.Label(disp_cell, text=f"Disparo\n{r['disparo']}",
                 font=FONT_BOLD, bg=bg, fg=fg,
                 justify="center").pack(expand=True)
        disp_cell.bind("<Button-1>", self._toggle)

        # ── Separador ──
        tk.Frame(self._hdr, bg=COLORS["border"], width=1).pack(side="left", fill="y")

        # ── Columna ARCHIVO 1 (centro) ──
        a1_cell = tk.Frame(self._hdr, bg=bg)
        a1_cell.pack(side="left", fill="both", expand=True)
        self._a1_lbl = tk.Label(a1_cell, font=FONT_SM, bg=bg, fg=fg,
                                anchor="center", justify="center")
        self._a1_lbl.pack(expand=True, fill="both", padx=6, pady=6)
        a1_cell.bind("<Button-1>", self._toggle)
        self._a1_lbl.bind("<Button-1>", self._toggle)

        # ── Separador ──
        tk.Frame(self._hdr, bg=COLORS["border"], width=1).pack(side="left", fill="y")

        # ── Columna ARCHIVO 2 (derecha) ──
        a2_cell = tk.Frame(self._hdr, bg=bg)
        a2_cell.pack(side="left", fill="both", expand=True)
        self._a2_lbl = tk.Label(a2_cell, font=FONT_SM, bg=bg, fg=fg,
                                anchor="center", justify="center")
        self._a2_lbl.pack(expand=True, fill="both", padx=6, pady=6)
        a2_cell.bind("<Button-1>", self._toggle)
        self._a2_lbl.bind("<Button-1>", self._toggle)

        # ── Flecha (solo si tiene detalle) ──
        if r["status"] == "diff":
            self._arrow = tk.Label(self._hdr, text=" ▾ ", font=FONT_SM,
                                   bg=bg, fg=fg)
            self._arrow.pack(side="right", padx=4)
            self._arrow.bind("<Button-1>", self._toggle)

        self._fill_cells()

    def _fill_cells(self):
        r = self._r
        s = r["status"]

        if s == "ok":
            n1 = r.get("n_lineas1", "—")
            self._a1_lbl.configure(text=f"{n1} líneas")
            self._a2_lbl.configure(text=f"{r.get('n_lineas2','—')} líneas")

        elif s == "only1":
            self._a1_lbl.configure(text=f"{r.get('n_lineas1','—')} líneas\n(presente)")
            self._a2_lbl.configure(text="—\n(ausente)")

        elif s == "only2":
            self._a1_lbl.configure(text="—\n(ausente)")
            self._a2_lbl.configure(text=f"{r.get('n_lineas2','—')} líneas\n(presente)")

        else:  # diff
            ndiff = sum(1 for lr in r["lineas"] if lr["status"] != "ok")
            nok   = sum(1 for lr in r["lineas"] if lr["status"] == "ok")
            self._a1_lbl.configure(
                text=f"{r.get('n_lineas1','—')} líneas\n{nok} ok · {ndiff} con difs.")
            self._a2_lbl.configure(
                text=f"{r.get('n_lineas2','—')} líneas\n{nok} ok · {ndiff} con difs.")

    def _toggle(self, _=None):
        if self._r["status"] != "diff":
            return
        self._open = not self._open
        if self._open:
            self._arrow.configure(text=" ▴ ")
            self._show_detail()
        else:
            self._arrow.configure(text=" ▾ ")
            if self._detail:
                self._detail.destroy()
                self._detail = None

    def _show_detail(self):
        """Panel de detalle: tabla de líneas con 3 columnas."""
        self._detail = tk.Frame(self, bg=COLORS["surface"],
                                highlightthickness=1,
                                highlightbackground=COLORS["border"])
        self._detail.pack(fill="x", padx=2, pady=(0, 4))

        # Cabecera de la tabla interna
        hdr = tk.Frame(self._detail, bg=COLORS["col_hdr"])
        hdr.pack(fill="x")
        for txt, w, side in [
            ("Línea",     100, "left"),
            ("Archivo 1", 0,   "left"),
            ("Archivo 2", 0,   "left"),
        ]:
            kw = {"width": w} if w else {}
            lbl = tk.Label(hdr, text=txt, font=FONT_BOLD,
                           bg=COLORS["col_hdr"], fg=COLORS["col_hdr_txt"],
                           anchor="center", padx=6, pady=4, **kw)
            lbl.pack(side=side, fill="both", expand=(w == 0))

        # Filas de líneas
        for i, lr in enumerate(self._r["lineas"]):
            s = lr["status"]
            row_bg = COLORS.get(STATUS_META.get(s, STATUS_META["diff"])[1], COLORS["surface"])
            row_fg = COLORS.get(STATUS_META.get(s, STATUS_META["diff"])[2], COLORS["text"])
            alt_bg = self._lighten(row_bg) if i % 2 == 0 else row_bg

            row = tk.Frame(self._detail, bg=alt_bg)
            row.pack(fill="x")

            # Celda línea
            lin_cell = tk.Frame(row, bg=alt_bg, width=100)
            lin_cell.pack(side="left", fill="y")
            lin_cell.pack_propagate(False)
            tk.Label(lin_cell, text=lr["linea"], font=FONT_MONO,
                     bg=alt_bg, fg=row_fg, anchor="center").pack(
                expand=True, fill="both", pady=3)

            tk.Frame(row, bg=COLORS["border"], width=1).pack(side="left", fill="y")

            # Celdas de contenido por archivo
            def fmt_cell(linea_r, arch):
                if arch == 1:
                    s_ = linea_r["status"]
                    if s_ == "only2": return "—"
                    lst = linea_r["estacas1"]
                else:
                    s_ = linea_r["status"]
                    if s_ == "only1": return "—"
                    lst = linea_r["estacas2"]
                if not lst: return "—"
                def _n(x): return float(x) if x.lstrip("-").replace(".","").isdigit() else 0
                ini = min((e[0] for e in lst), key=_n)
                fin = max((e[1] for e in lst), key=_n)
                return f"{ini} → {fin}" if len(lst) == 1 else f"{ini} → {fin}  ({len(lst)} rangos)"

            for arch in (1, 2):
                cell_txt = fmt_cell(lr, arch)
                # Marcar diferencias de estacas
                extra = ""
                if s == "diff":
                    missing = [ed for ed in lr["estaca_diffs"]
                               if (arch == 1 and ed["tipo"] == "solo_arch2") or
                                  (arch == 2 and ed["tipo"] == "solo_arch1")]
                    if missing:
                        extra = "\n⚠ " + "  ".join(f"{e['ini']}→{e['fin']}" for e in missing[:3])
                        if len(missing) > 3:
                            extra += f" +{len(missing)-3}"

                a_cell = tk.Frame(row, bg=alt_bg)
                a_cell.pack(side="left", fill="both", expand=True)
                tk.Label(a_cell, text=cell_txt + extra,
                         font=FONT_MONO, bg=alt_bg, fg=row_fg,
                         anchor="center", justify="center",
                         padx=6, pady=3).pack(fill="both", expand=True)
                if arch == 1:
                    tk.Frame(row, bg=COLORS["border"], width=1).pack(side="left", fill="y")

            # Separador entre filas
            tk.Frame(self._detail, bg=COLORS["border"], height=1).pack(fill="x")

    @staticmethod
    def _lighten(hex_color):
        """Aclara un color hexadecimal mezclándolo con blanco."""
        try:
            r = int(hex_color[1:3], 16)
            g = int(hex_color[3:5], 16)
            b = int(hex_color[5:7], 16)
            r = min(255, r + (255 - r) // 3)
            g = min(255, g + (255 - g) // 3)
            b = min(255, b + (255 - b) // 3)
            return f"#{r:02x}{g:02x}{b:02x}"
        except Exception:
            return hex_color


# ─── Aplicación principal ─────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Comparador SPS / XPS")
        self.configure(bg=COLORS["bg"])
        self.resizable(True, True)
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w  = min(max(int(sw * 0.90), 900), sw)
        h  = min(max(int(sh * 0.88), 640), sh)
        x  = (sw - w) // 2
        y  = max((sh - h) // 2 - 30, 0)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(900, 640)
        self._result = None
        self._disp_frames = []
        self._build()

    def _build(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        # ── HEADER ──────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=COLORS["surface"],
                       highlightthickness=1,
                       highlightbackground=COLORS["border"])
        hdr.pack(fill="x")
        tk.Label(hdr, text="Comparador de scripts de disparo",
                 font=FONT_H1, bg=COLORS["surface"],
                 fg=COLORS["text"]).pack(side="left", padx=20, pady=12)
        tk.Label(hdr, text="Disparo  ›  Línea  ›  Estacas  ·  Creado por:   Pedro R. Bectto Fayad",
                 font=FONT_SM, bg=COLORS["surface"],
                 fg=COLORS["muted"]).pack(side="left", pady=12)

        # ── SELECCIÓN ───────────────────────────────────────────────────────
        sel = tk.Frame(self, bg=COLORS["bg"])
        sel.pack(fill="x", padx=20, pady=(14, 6))
        self._f1 = FileSelector(sel, "Archivo 1")
        self._f1.pack(side="left", expand=True, fill="both", padx=(0, 8))
        self._f2 = FileSelector(sel, "Archivo 2")
        self._f2.pack(side="left", expand=True, fill="both")

        # ── BOTÓN COMPARAR ───────────────────────────────────────────────────
        btn_f = tk.Frame(self, bg=COLORS["bg"])
        btn_f.pack(fill="x", padx=20, pady=4)
        self._btn = tk.Button(btn_f, text="  ⇄  Comparar archivos  ",
                              font=FONT_BOLD, cursor="hand2",
                              bg=COLORS["accent"], fg="#FFF",
                              relief="flat", bd=0, padx=16, pady=8,
                              activebackground="#333",
                              command=self._on_compare)
        self._btn.pack(side="right")

        # ── MÉTRICAS ────────────────────────────────────────────────────────
        met = tk.Frame(self, bg=COLORS["bg"])
        met.pack(fill="x", padx=20, pady=(6, 0))
        self._m_d1   = MetricCard(met, "DISPAROS ARCHIVO 1")
        self._m_d2   = MetricCard(met, "DISPAROS ARCHIVO 2")
        self._m_diff = MetricCard(met, "DISPAROS CON DIFERENCIAS DE LÍNEAS/ESTACAS")
        self._m_miss = MetricCard(met, "DISPAROS FALTANTES")
        for m in (self._m_d1, self._m_d2, self._m_diff, self._m_miss):
            m.pack(side="left", expand=True, fill="x", padx=(0, 8), pady=4)

        # ── AVISO EXTRA ─────────────────────────────────────────────────────
        self._extra_lbl = tk.Label(self, text="", font=FONT_SM,
                                   bg="#FFF3CD", fg=COLORS["diff_txt"],
                                   anchor="w", padx=12)

        # ── LEYENDA ─────────────────────────────────────────────────────────
        leg = tk.Frame(self, bg=COLORS["bg"])
        leg.pack(fill="x", padx=20, pady=(6, 0))
        for bg_key, fg_key, txt in [
            ("ok_row",    "ok_txt",    "■ Idéntico"),
            ("diff_row",  "diff_txt",  "■ Con diferencias"),
            ("only1_row", "only1_txt", "■ Solo en Archivo 1"),
            ("only2_row", "only2_txt", "■ Solo en Archivo 2"),
        ]:
            f = tk.Frame(leg, bg=COLORS[bg_key],
                         highlightthickness=1,
                         highlightbackground=COLORS["border"])
            f.pack(side="left", padx=(0, 6), pady=4)
            tk.Label(f, text=txt, font=FONT_SM,
                     fg=COLORS[fg_key], bg=COLORS[bg_key],
                     padx=8, pady=3).pack()

        # ── CABECERA DE COLUMNAS ─────────────────────────────────────────────
        self._col_hdr = tk.Frame(self, bg=COLORS["col_hdr"],
                                 highlightthickness=1,
                                 highlightbackground=COLORS["border"])
        self._col_hdr.pack(fill="x", padx=20, pady=(4, 0))

        # ── FILTROS + BOTONES EXPORTAR ───────────────────────────────────────
        filt = tk.Frame(self, bg=COLORS["bg"])
        filt.pack(fill="x", padx=20, pady=(4, 2))

        # Botones exportar (derecha)
        for txt, cmd, bg_c, fg_c in [
            ("Exportar Comparativa", self._export_comparativa, COLORS["ok_row"],   COLORS["ok_txt"]),
            ("Exportar TXT",         self._export_txt,         COLORS["col_hdr"],  COLORS["text"]),
            ("Exportar Excel (.xlsx)", self._export_xlsx,       COLORS["ok_row"],   COLORS["ok_txt"]),
        ]:
            tk.Button(filt, text=txt, font=FONT_SM, cursor="hand2",
                      relief="flat", bd=1, padx=10, pady=4,
                      bg=bg_c, fg=fg_c,
                      activebackground=COLORS["border"],
                      command=cmd).pack(side="right", padx=(4, 0))

        # Separador visual
        tk.Frame(filt, bg=COLORS["border"], width=1).pack(side="right", fill="y", padx=6)

        # Búsqueda
        tk.Entry(filt, textvariable=tk.StringVar(), font=FONT_SM, width=16,
                 relief="flat", bd=1,
                 highlightthickness=1,
                 highlightbackground=COLORS["border"]).pack(side="right")
        self._svar = tk.StringVar()
        self._svar.trace_add("write", lambda *_: self._apply_filter())
        # re-bind entry to correct var
        filt.winfo_children()[-1].configure(textvariable=self._svar)

        tk.Label(filt, text="Buscar:", font=FONT_SM,
                 bg=COLORS["bg"], fg=COLORS["muted"]).pack(side="right", padx=(4, 2))

        # Filtros radio (izquierda)
        tk.Label(filt, text="Mostrar:", font=FONT_SM,
                 bg=COLORS["bg"], fg=COLORS["muted"]).pack(side="left")
        self._fvar = tk.StringVar(value="all")
        for lbl, val in [("Todos","all"), ("Con diferencias","diff"),
                          ("Idénticos","ok"), ("Faltantes","missing")]:
            tk.Radiobutton(filt, text=lbl, value=val, variable=self._fvar,
                           font=FONT_SM, bg=COLORS["bg"], fg=COLORS["text"],
                           activebackground=COLORS["bg"],
                           command=self._apply_filter).pack(side="left", padx=5)

        # ── LISTA SCROLLABLE ─────────────────────────────────────────────────
        outer = tk.Frame(self, bg=COLORS["bg"])
        outer.pack(fill="both", expand=True, padx=20, pady=(0, 4))

        self._canvas = tk.Canvas(outer, bg=COLORS["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)

        self._inner = tk.Frame(self._canvas, bg=COLORS["bg"])
        self._win = self._canvas.create_window((0, 0), window=self._inner, anchor="nw")
        self._inner.bind("<Configure>",
                         lambda e: self._canvas.configure(
                             scrollregion=self._canvas.bbox("all")))
        self._canvas.bind("<Configure>",
                          lambda e: self._canvas.itemconfig(self._win, width=e.width))
        self._canvas.bind_all("<MouseWheel>",
                              lambda e: self._canvas.yview_scroll(
                                  int(-1*(e.delta/120)), "units"))

        tk.Label(self._inner,
                 text="Carga dos archivos y pulsa «Comparar».",
                 font=FONT_SM, bg=COLORS["bg"],
                 fg=COLORS["muted"]).pack(pady=30)

        # ── BARRA INFERIOR ───────────────────────────────────────────────────
        bot = tk.Frame(self, bg=COLORS["surface"],
                       highlightthickness=1,
                       highlightbackground=COLORS["border"])
        bot.pack(fill="x", side="bottom")
        self._status = tk.Label(bot, text="Listo.", font=FONT_SM,
                                bg=COLORS["surface"], fg=COLORS["muted"], anchor="w")
        self._status.pack(side="left", padx=12, pady=5)

    def _build_col_headers(self):
        """Dibuja la cabecera de 3 columnas una vez que hay resultado."""
        for w in self._col_hdr.winfo_children():
            w.destroy()

        # Columna disparo
        d_cell = tk.Frame(self._col_hdr, bg=COLORS["col_hdr"], width=COL_DISP)
        d_cell.pack(side="left", fill="y")
        d_cell.pack_propagate(False)
        tk.Label(d_cell, text="DISPARO", font=FONT_BOLD,
                 bg=COLORS["col_hdr"], fg=COLORS["col_hdr_txt"],
                 anchor="center").pack(expand=True, fill="both", pady=5)

        tk.Frame(self._col_hdr, bg=COLORS["border"], width=1).pack(side="left", fill="y")

        r = self._result
        for txt in (f"ARCHIVO 1  ·  {r['name1']}", f"ARCHIVO 2  ·  {r['name2']}"):
            cell = tk.Frame(self._col_hdr, bg=COLORS["col_hdr"])
            cell.pack(side="left", fill="both", expand=True)
            tk.Label(cell, text=txt, font=FONT_BOLD,
                     bg=COLORS["col_hdr"], fg=COLORS["col_hdr_txt"],
                     anchor="center").pack(fill="both", pady=5)
            tk.Frame(self._col_hdr, bg=COLORS["border"], width=1).pack(side="left", fill="y")

    # ── acciones ─────────────────────────────────────────────────────────────

    def _on_compare(self):
        p1, p2 = self._f1.get(), self._f2.get()
        if not p1 or not p2:
            messagebox.showwarning("Faltan archivos",
                                   "Selecciona los dos archivos antes de comparar.")
            return
        self._status.configure(text="Comparando…")
        self.update()
        try:
            self._result = run_comparison(p1, p2)
        except Exception as ex:
            messagebox.showerror("Error", str(ex))
            self._status.configure(text="Error.")
            return
        self._render_results()

    def _render_results(self):
        r = self._result
        n1, n2 = r["n_disparos1"], r["n_disparos2"]
        ndiff  = len(r["disparos_diff"])
        nmiss  = len(r["disparos_solo1"]) + len(r["disparos_solo2"])

        self._m_d1.set(n1)
        self._m_d2.set(n2, COLORS["diff_txt"] if n1 != n2 else COLORS["text"])
        self._m_diff.set(ndiff, COLORS["diff_txt"] if ndiff else COLORS["ok_txt"])
        self._m_miss.set(nmiss, COLORS["only1_txt"] if nmiss else COLORS["ok_txt"])

        if n1 != n2:
            diff  = abs(n1 - n2)
            which = "Archivo 1" if n1 > n2 else "Archivo 2"
            self._extra_lbl.configure(
                text=f"  ⚠  {which} tiene {diff} disparo(s) adicional(es) sin par.")
            self._extra_lbl.pack(fill="x", padx=20, pady=(0, 2))
        else:
            self._extra_lbl.pack_forget()

        # Mostrar cabecera de columnas
        self._build_col_headers()

        # Construir filas
        for w in self._inner.winfo_children():
            w.destroy()
        self._disp_frames = []

        for shot in r["results"]:
            row = DisparoRow(self._inner, shot)
            row.pack(fill="x", pady=(0, 2))
            self._disp_frames.append((row, shot))

        self._apply_filter()

        total = len(r["results"])
        ok    = len(r["disparos_ok"])
        self._status.configure(
            text=(f"✔  {total} disparos  ·  {ok} idénticos  ·  "
                  f"{ndiff} con diferencias  ·  {nmiss} faltantes"))

    def _apply_filter(self):
        filt   = self._fvar.get()
        search = self._svar.get().strip().lower()
        for row, s in self._disp_frames:
            show = True
            if filt == "diff"    and s["status"] != "diff":              show = False
            if filt == "ok"      and s["status"] != "ok":                show = False
            if filt == "missing" and s["status"] not in ("only1","only2"): show = False
            if show and search and search not in s["disparo"].lower():    show = False
            if show: row.pack(fill="x", pady=(0, 2))
            else:    row.pack_forget()

    def _export_txt(self):
        if not self._result:
            messagebox.showwarning("Sin datos", "Ejecuta una comparación primero.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt")],
            initialfile="reporte_diferencias.txt")
        if path:
            export_txt(self._result, path)
            self._status.configure(text=f"TXT guardado: {os.path.basename(path)}")

    def _export_xlsx(self):
        if not self._result:
            messagebox.showwarning("Sin datos", "Ejecuta una comparación primero.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="comparativa.xlsx")
        if path:
            try:
                export_xlsx(self._result, path)
                self._status.configure(
                    text=f"Excel guardado: {os.path.basename(path)}  "
                         f"(3 hojas: Resumen · Comparativa completa · Solo diferencias)")
            except ImportError as e:
                messagebox.showerror("Módulo faltante", str(e))

    def _export_comparativa(self):
        if not self._result:
            messagebox.showwarning("Sin datos", "Ejecuta una comparación primero.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt")],
            initialfile="comparativa_completa.txt")
        if path:
            export_comparativa(self._result, path)
            self._status.configure(text=f"Comparativa guardada: {os.path.basename(path)}")


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()